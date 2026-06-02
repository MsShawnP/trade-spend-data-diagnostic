"""End-to-end validation of the trade spend diagnostic workbook."""

import contextlib
import os
import sqlite3
import sys
from dataclasses import dataclass, field
from pathlib import Path

os.environ.setdefault("PYTHONIOENCODING", "utf-8")
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

OUTPUT = Path(__file__).resolve().parent / "output" / "trade_spend_diagnostic.xlsx"
DB = Path(__file__).resolve().parent / "cinderhaven-data" / "data" / "cinderhaven_product_master.db"


@dataclass
class Results:
    passed: int = 0
    failed: int = 0
    details: list[str] = field(default_factory=list)

    def check(self, name: str, condition: bool, detail: str = "") -> None:
        if condition:
            self.passed += 1
            print(f"  PASS: {name}")
        else:
            self.failed += 1
            print(f"  FAIL: {name}")
            if detail:
                print(f"        {detail}")
                self.details.append(f"{name}: {detail}")


def approx(a: float | int | None, b: float | int | None, tolerance: float = 0.005) -> bool:
    if a is None or b is None:
        return False
    if b == 0:
        return a == 0
    return abs(a - b) / abs(b) < tolerance


def main() -> bool:
    print(f"Validating: {OUTPUT}")
    print(f"Database:   {DB}")
    print()

    r = Results()
    wb = load_workbook(OUTPUT, data_only=False)

    with contextlib.closing(sqlite3.connect(DB)) as conn:
        # === TAB STRUCTURE ===
        print("=== Tab Structure ===")
        expected_tabs = [
            "Executive Pulse", "Leak Diagnostic", "Promo Efficacy",
            "Retailer Risk", "Deduction Ledger", "Deduction Code Crosswalk",
            "Methodology & Logic",
        ]
        r.check("Tab count is 7", len(wb.sheetnames) == 7,
                f"Got {len(wb.sheetnames)}: {wb.sheetnames}")
        r.check("No default 'Sheet' tab", "Sheet" not in wb.sheetnames)
        r.check("Tab order correct", wb.sheetnames == expected_tabs,
                f"Got {wb.sheetnames}")
        r.check("Active sheet is Tab 1", wb.active.title == "Executive Pulse",
                f"Got {wb.active.title}")

        chicago_tabs = ["Executive Pulse", "Leak Diagnostic", "Promo Efficacy", "Retailer Risk"]
        for tab in chicago_tabs:
            color = wb[tab].sheet_properties.tabColor.rgb if wb[tab].sheet_properties.tabColor else ""
            r.check(f"{tab} tab color Chicago-20", "1f2e7a" in color.lower(), f"Got {color}")

        color = wb["Deduction Ledger"].sheet_properties.tabColor.rgb
        r.check("Deduction Ledger tab color HK-35", "158f75" in color.lower(), f"Got {color}")

        for tab in ["Deduction Code Crosswalk", "Methodology & Logic"]:
            color = wb[tab].sheet_properties.tabColor.rgb if wb[tab].sheet_properties.tabColor else ""
            r.check(f"{tab} tab color London-70", "b3b3b3" in color.lower(), f"Got {color}")

        # === LOCKED NUMBERS (Tab 1) ===
        print()
        print("=== Locked Numbers (Tab 1) ===")
        ws1 = wb["Executive Pulse"]

        revenue = ws1["D11"].value
        r.check("Revenue ≈ $24,616,311", approx(revenue, 24616311),
                f"Got ${revenue:,.0f}" if revenue else "Got None")

        structural = ws1["D12"].value
        r.check("Structural trade ≈ $4,311,076", approx(structural, 4311076),
                f"Got ${structural:,.0f}" if structural else "Got None")

        waste = ws1["D13"].value
        r.check("Operational waste ≈ $1,030,034", approx(waste, 1030034, 0.02),
                f"Got ${waste:,.0f}" if waste else "Got None")

        all_in_rate = ws1["B5"].value
        r.check("All-in trade rate ≈ 21.7%",
                approx(all_in_rate, 0.217, 0.01),
                f"Got {all_in_rate*100:.1f}%" if all_in_rate else "Got None")

        structural_rate = ws1["C5"].value
        r.check("Structural trade rate = 17.5%",
                approx(structural_rate, 0.1751, 0.005),
                f"Got {structural_rate*100:.1f}%" if structural_rate else "Got None")

        waste_rate = ws1["D5"].value
        r.check("Operational waste rate ≈ 4.2%",
                approx(waste_rate, 0.0418, 0.02),
                f"Got {waste_rate*100:.1f}%" if waste_rate else "Got None")

        # === RECOVERY RATE VALIDATION ===
        print()
        print("=== Recovery Rate ===")
        db_recovered = conn.execute(
            "SELECT COALESCE(SUM(recovered_amount), 0) FROM disputes"
        ).fetchone()[0]
        db_disputed = conn.execute(
            "SELECT COALESCE(SUM(d.amount), 0) FROM deductions d "
            "JOIN disputes dis ON dis.deduction_id = d.deduction_id"
        ).fetchone()[0]
        expected_recovery = db_recovered / db_disputed if db_disputed else 0
        wb_recovery = ws1["C25"].value
        r.check(f"Recovery rate ≈ {expected_recovery:.1%}",
                approx(wb_recovery, expected_recovery, 0.01),
                f"Got {wb_recovery:.1%}" if wb_recovery else "Got None")
        r.check(f"Disputes total > 3000",
                conn.execute("SELECT COUNT(*) FROM disputes").fetchone()[0] > 3000)

        # === DOUBLE-DIP CHECK (Tab 2) ===
        print()
        print("=== Double-Dip Check (Tab 2) ===")
        ws2 = wb["Leak Diagnostic"]

        dd_header_row = None
        for row in range(1, ws2.max_row + 1):
            val = ws2.cell(row=row, column=2).value
            if val and "Double-Dip" in str(val) and "Alert" in str(val):
                dd_header_row = row
                break

        dd_count = 0
        dd_total = 0.0
        if dd_header_row:
            for row in range(dd_header_row + 1, ws2.max_row + 1):
                cell = ws2.cell(row=row, column=2)
                if cell.value and str(cell.value).startswith("DED-"):
                    dd_count += 1
                    dd_total += ws2.cell(row=row, column=4).value or 0

        r.check("Found Double-Dip Alert section", dd_header_row is not None,
                "Could not find 'Double-Dip Alert' header")
        r.check("3 double-dip events", dd_count == 3, f"Got {dd_count}")
        r.check("Double-dip total ≈ $19,372", approx(dd_total, 19372, 0.02),
                f"Got ${dd_total:,.0f}")

        # === RETAILER TOTALS (Tab 4) ===
        print()
        print("=== Retailer Totals (Tab 4) ===")
        ws4 = wb["Retailer Risk"]

        retailer_rev_sum = 0
        retailer_struct_sum = 0
        for row in range(6, 17):
            rev = ws4.cell(row=row, column=3).value or 0
            struct = ws4.cell(row=row, column=5).value or 0
            retailer_rev_sum += rev
            retailer_struct_sum += struct

        r.check("Tab 4 revenue sums to total",
                approx(retailer_rev_sum, revenue),
                f"Got ${retailer_rev_sum:,.0f} vs ${revenue:,.0f}" if revenue else "Revenue is None")
        r.check("Tab 4 structural sums to total",
                approx(retailer_struct_sum, structural),
                f"Got ${retailer_struct_sum:,.0f} vs ${structural:,.0f}" if structural else "Structural is None")

        # === DEDUCTION COUNT (Tab 5) ===
        print()
        print("=== Deduction Count (Tab 5) ===")
        ws5 = wb["Deduction Ledger"]

        weeks = conn.execute(
            "SELECT DISTINCT week_ending FROM scan_data ORDER BY week_ending DESC LIMIT 52"
        ).fetchall()
        max_scan = weeks[0][0]
        db_count = conn.execute("""
            SELECT COUNT(*) FROM deductions
            WHERE deduction_date > date(?, '-365 days') AND deduction_date <= ?
        """, (max_scan, max_scan)).fetchone()[0]

        ledger_rows = 0
        for row in range(5, ws5.max_row + 1):
            if ws5.cell(row=row, column=1).value:
                ledger_rows += 1
            else:
                break

        r.check(f"Tab 5 rows match DB ({db_count})", ledger_rows == db_count,
                f"Tab 5 has {ledger_rows} rows, DB has {db_count}")

        # === CROSSWALK COMPLETENESS (Tab 6) ===
        print()
        print("=== Crosswalk Completeness (Tab 6) ===")
        ws6 = wb["Deduction Code Crosswalk"]

        tab5_retailers = set()
        for row in range(5, 5 + ledger_rows):
            ret = ws5.cell(row=row, column=3).value
            if ret:
                tab5_retailers.add(ret.lower().replace(" ", "_"))

        tab6_retailers = set()
        for row in range(5, ws6.max_row + 1):
            ret = ws6.cell(row=row, column=1).value
            if ret:
                tab6_retailers.add(ret.lower().replace(" ", "_"))

        missing = tab5_retailers - tab6_retailers
        r.check("Tab 6 covers all Tab 5 retailers", len(missing) == 0,
                f"Missing: {missing}" if missing else "")

        # === PROMO EFFICACY CONTENT (Tab 3) ===
        print()
        print("=== Promo Efficacy Content (Tab 3) ===")
        ws3 = wb["Promo Efficacy"]

        tab3_header = ws3["B1"].value or ws3["A1"].value
        r.check("Tab 3 header present",
                tab3_header is not None and "Promo" in str(tab3_header),
                f"Got B1={ws3['B1'].value}, A1={ws3['A1'].value}")

        promo_tables = [t.displayName for t in ws3.tables.values()]
        r.check("Tab 3 has promo table", "tbl_PromoEfficacy" in promo_tables,
                f"Found: {promo_tables}")

        promo_data_rows = 0
        for row in range(10, ws3.max_row + 1):
            if ws3.cell(row=row, column=2).value:
                promo_data_rows += 1
        r.check("Tab 3 has promo data rows", promo_data_rows > 0,
                f"Found {promo_data_rows} data rows")

        # === METHODOLOGY CONTENT (Tab 7) ===
        print()
        print("=== Methodology Content (Tab 7) ===")
        ws7 = wb["Methodology & Logic"]

        r.check("Tab 7 header present",
                ws7["A1"].value is not None and "Methodology" in str(ws7["A1"].value),
                f"Got A1={ws7['A1'].value}")

        methodology_sections = 0
        for row in range(1, ws7.max_row + 1):
            val = ws7.cell(row=row, column=1).value
            if val and str(val).startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.")):
                methodology_sections += 1
        r.check("Tab 7 has all 7 sections", methodology_sections == 7,
                f"Found {methodology_sections} sections")

        # === CROSS-TAB CONSISTENCY ===
        print()
        print("=== Cross-Tab Consistency ===")

        tab2_cat_total = 0
        for row in range(6, 14):
            amt = ws2.cell(row=row, column=4).value
            if isinstance(amt, (int, float)):
                tab2_cat_total += amt

        r.check("Tab 2 categories sum ≈ Tab 1 waste",
                approx(tab2_cat_total, waste),
                f"Tab 2 sum: ${tab2_cat_total:,.0f} vs Tab 1: ${waste:,.0f}" if waste else "Waste is None")

        r.check("Tab 4 rev sum = Tab 1 revenue",
                approx(retailer_rev_sum, revenue),
                f"${retailer_rev_sum:,.0f} vs ${revenue:,.0f}" if revenue else "Revenue is None")

        # === NO ERRORS (full scan, no row cap) ===
        print()
        print("=== Error Scan ===")
        error_values = {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!"}
        errors_found = []
        for tab_name in wb.sheetnames:
            ws = wb[tab_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
                for cell in row:
                    if cell.value and str(cell.value) in error_values:
                        errors_found.append(f"{tab_name}!{cell.coordinate}: {cell.value}")

        r.check("No #N/A, #REF, #VALUE errors", len(errors_found) == 0,
                f"Found: {errors_found[:5]}" if errors_found else "")

        # === NAMED RANGES ===
        print()
        print("=== Named Ranges ===")
        expected_names = [
            "AllInTradeRate", "StructuralTradeRate", "OperationalWasteRate",
            "TotalRevenue", "StructuralTrade", "OperationalWaste",
            "AllInTradeCost", "RecoveryRate",
        ]
        defined = [dn.name for dn in wb.defined_names.values()]
        for name in expected_names:
            r.check(f"Named range '{name}' defined", name in defined,
                    f"Defined names: {defined}" if name not in defined else "")

        no_kpi_aliases = [n for n in defined if n.startswith("KPI_")]
        r.check("No stale KPI_ alias ranges", len(no_kpi_aliases) == 0,
                f"Found stale aliases: {no_kpi_aliases}" if no_kpi_aliases else "")

        # === EXCEL TABLES ===
        print()
        print("=== Excel Tables ===")
        expected_tables = {
            "Executive Pulse": ["tbl_AddressableImprovement", "tbl_ResponsibilityMatrix"],
            "Leak Diagnostic": ["tbl_WasteByCategory", "tbl_DoubleDips"],
            "Promo Efficacy": ["tbl_PromoEfficacy"],
            "Retailer Risk": ["tbl_RetailerPnL", "tbl_ConcentrationRisk"],
            "Deduction Ledger": ["tbl_DeductionLedger"],
            "Deduction Code Crosswalk": ["tbl_CodeCrosswalk"],
        }
        for tab_name, table_names in expected_tables.items():
            ws = wb[tab_name]
            actual_tables = [t.displayName for t in ws.tables.values()]
            for tbl in table_names:
                r.check(f"Table '{tbl}' in {tab_name}", tbl in actual_tables,
                        f"Found tables: {actual_tables}" if tbl not in actual_tables else "")

        # === DATA VALIDATION ===
        print()
        print("=== Data Validation ===")
        ws2_dvs = ws2.data_validations.dataValidation
        r.check("Tab 2 has data validation", len(ws2_dvs) > 0,
                f"Found {len(ws2_dvs)} validations")

        ws3_dvs = ws3.data_validations.dataValidation
        r.check("Tab 3 has data validation", len(ws3_dvs) > 0,
                f"Found {len(ws3_dvs)} validations")

        ws4_dvs = ws4.data_validations.dataValidation
        r.check("Tab 4 has data validation", len(ws4_dvs) > 0,
                f"Found {len(ws4_dvs)} validations")

        # === CONDITIONAL FORMATTING ===
        print()
        print("=== Conditional Formatting ===")
        ws1_cf = list(ws1.conditional_formatting)
        r.check("Tab 1 has conditional formatting (data bars)", len(ws1_cf) >= 4,
                f"Found {len(ws1_cf)} rules")

        ws2_cf = list(ws2.conditional_formatting)
        r.check("Tab 2 has conditional formatting", len(ws2_cf) > 0,
                f"Found {len(ws2_cf)} rules")

        ws3_cf = list(ws3.conditional_formatting)
        r.check("Tab 3 has conditional formatting (ROI + quality)", len(ws3_cf) >= 2,
                f"Found {len(ws3_cf)} ranges")

        ws4_cf = list(ws4.conditional_formatting)
        r.check("Tab 4 has conditional formatting (margin)", len(ws4_cf) > 0,
                f"Found {len(ws4_cf)} ranges")

        ws6_cf = list(ws6.conditional_formatting)
        r.check("Tab 6 has conditional formatting (verified/inferred)", len(ws6_cf) > 0,
                f"Found {len(ws6_cf)} ranges")

    # === SUMMARY ===
    print()
    print("=" * 50)
    print(f"RESULTS: {r.passed} passed, {r.failed} failed")
    if r.failed == 0:
        print("All checks passed.")
    else:
        print("SOME CHECKS FAILED — see details above.")
    print("=" * 50)

    return r.failed == 0


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
