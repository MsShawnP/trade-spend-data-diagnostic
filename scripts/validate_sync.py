"""Cross-validate workbook KPIs against computed CSV values.

Ensures the workbook (generated from CSVs) matches the CSV source data.
This is the contract check: CSVs are the single source of truth.

Usage:
    python -m scripts.validate_sync
"""

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "output" / "trade_spend_diagnostic.xlsx"
DATA_DIR = ROOT / "powerbi" / "data"

PASS = 0
FAIL = 0


def check(name: str, condition: bool, detail: str = ""):
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"  PASS: {name}")
    else:
        FAIL += 1
        print(f"  FAIL: {name}")
        if detail:
            print(f"        {detail}")


def approx(a, b, tolerance=0.005):
    if b == 0:
        return a == 0
    return abs(a - b) / abs(b) < tolerance


def main():
    global PASS, FAIL

    print(f"Workbook: {OUTPUT}")
    print(f"CSV dir:  {DATA_DIR}")
    print()

    wb = load_workbook(OUTPUT, data_only=False)

    # Load CSV KPIs
    with open(DATA_DIR / "computed_kpis.csv", encoding="utf-8") as f:
        kpi = next(csv.DictReader(f))

    csv_revenue = float(kpi["revenue"])
    csv_structural = float(kpi["structural_trade"])
    csv_waste = float(kpi["operational_waste"])
    csv_all_in_rate = float(kpi["all_in_trade_rate"])
    csv_struct_rate = float(kpi["structural_trade_rate"])
    csv_waste_rate = float(kpi["waste_rate"])
    csv_recovered = float(kpi["total_recovered"])
    csv_recovery_rate = float(kpi["recovery_rate"])

    # ── Tab 1 KPIs vs CSV ──
    print("=== Tab 1 KPIs vs CSV ===")
    ws1 = wb["Executive Pulse"]

    check("Revenue matches CSV",
          approx(ws1["D11"].value, csv_revenue),
          f"Workbook: {ws1['D11'].value:,.0f} vs CSV: {csv_revenue:,.0f}")

    check("Structural trade matches CSV",
          approx(ws1["D12"].value, csv_structural),
          f"Workbook: {ws1['D12'].value:,.0f} vs CSV: {csv_structural:,.0f}")

    check("Operational waste matches CSV",
          approx(ws1["D13"].value, csv_waste),
          f"Workbook: {ws1['D13'].value:,.0f} vs CSV: {csv_waste:,.0f}")

    check("All-in trade rate matches CSV",
          approx(ws1["B5"].value, csv_all_in_rate),
          f"Workbook: {ws1['B5'].value:.4f} vs CSV: {csv_all_in_rate:.4f}")

    check("Structural trade rate matches CSV",
          approx(ws1["C5"].value, csv_struct_rate),
          f"Workbook: {ws1['C5'].value:.4f} vs CSV: {csv_struct_rate:.4f}")

    check("Operational waste rate matches CSV",
          approx(ws1["D5"].value, csv_waste_rate),
          f"Workbook: {ws1['D5'].value:.4f} vs CSV: {csv_waste_rate:.4f}")

    # ── Tab 2 waste categories vs CSV ──
    print()
    print("=== Tab 2 Categories vs CSV ===")
    ws2 = wb["Leak Diagnostic"]

    csv_categories = []
    with open(DATA_DIR / "waste_by_category.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            csv_categories.append(row)

    # Sum Tab 2 category amounts (rows 6-13, col D)
    wb_cat_total = 0
    for r in range(6, 6 + len(csv_categories)):
        amt = ws2.cell(row=r, column=4).value
        if isinstance(amt, (int, float)):
            wb_cat_total += amt

    csv_cat_total = sum(float(c["total_amount"]) for c in csv_categories)
    check("Category totals match CSV",
          approx(wb_cat_total, csv_cat_total),
          f"Workbook: ${wb_cat_total:,.0f} vs CSV: ${csv_cat_total:,.0f}")

    # ── Tab 2 double-dips vs CSV ──
    print()
    print("=== Tab 2 Double-Dips vs CSV ===")

    csv_dd = []
    with open(DATA_DIR / "double_dips.csv", encoding="utf-8") as f:
        csv_dd = list(csv.DictReader(f))

    csv_dd_total = sum(float(d["amount"]) for d in csv_dd)

    # Find double-dip section in workbook
    wb_dd_count = 0
    wb_dd_total = 0
    for r in range(1, ws2.max_row + 1):
        cell = ws2.cell(row=r, column=2)
        if cell.value and str(cell.value).startswith("DED-"):
            if r > 15:  # past the category table
                wb_dd_count += 1
                wb_dd_total += ws2.cell(row=r, column=4).value or 0

    check("Double-dip count matches CSV",
          wb_dd_count == len(csv_dd),
          f"Workbook: {wb_dd_count} vs CSV: {len(csv_dd)}")

    check("Double-dip total matches CSV",
          approx(wb_dd_total, csv_dd_total),
          f"Workbook: ${wb_dd_total:,.0f} vs CSV: ${csv_dd_total:,.0f}")

    # ── Tab 4 retailer totals vs CSV ──
    print()
    print("=== Tab 4 Retailers vs CSV ===")

    csv_retailers = []
    with open(DATA_DIR / "dim_retailer.csv", encoding="utf-8") as f:
        csv_retailers = list(csv.DictReader(f))

    csv_rev_total = sum(float(r["revenue"]) for r in csv_retailers)
    csv_struct_total = sum(float(r["structural_trade_dollars"]) for r in csv_retailers)

    ws4 = wb["Retailer Risk"]
    wb_rev_sum = 0
    wb_struct_sum = 0
    for r in range(6, 17):
        wb_rev_sum += ws4.cell(row=r, column=3).value or 0
        wb_struct_sum += ws4.cell(row=r, column=5).value or 0

    check("Tab 4 revenue sums match CSV total",
          approx(wb_rev_sum, csv_rev_total),
          f"Workbook: ${wb_rev_sum:,.0f} vs CSV: ${csv_rev_total:,.0f}")

    check("Tab 4 structural sums match CSV total",
          approx(wb_struct_sum, csv_struct_total),
          f"Workbook: ${wb_struct_sum:,.0f} vs CSV: ${csv_struct_total:,.0f}")

    # ── Tab 5 row count vs CSV ──
    print()
    print("=== Tab 5 Ledger vs CSV ===")
    ws5 = wb["Deduction Ledger"]

    csv_trailing = 0
    with open(DATA_DIR / "fact_deductions.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row["in_trailing_window"] == "1":
                csv_trailing += 1

    wb_ledger_rows = 0
    for r in range(5, ws5.max_row + 1):
        if ws5.cell(row=r, column=1).value:
            wb_ledger_rows += 1
        else:
            break

    check("Tab 5 row count matches CSV trailing window",
          wb_ledger_rows == csv_trailing,
          f"Workbook: {wb_ledger_rows} vs CSV: {csv_trailing}")

    # ── Tab 6 code count vs CSV ──
    print()
    print("=== Tab 6 Crosswalk vs CSV ===")
    ws6 = wb["Deduction Code Crosswalk"]

    csv_codes = 0
    with open(DATA_DIR / "deduction_codes.csv", encoding="utf-8") as f:
        csv_codes = sum(1 for _ in csv.DictReader(f))

    wb_code_rows = 0
    for r in range(5, ws6.max_row + 1):
        if ws6.cell(row=r, column=1).value:
            wb_code_rows += 1
        else:
            break

    check("Tab 6 code count matches CSV",
          wb_code_rows == csv_codes,
          f"Workbook: {wb_code_rows} vs CSV: {csv_codes}")

    # ── Summary ──
    print()
    print("=" * 50)
    print(f"RESULTS: {PASS} passed, {FAIL} failed")
    if FAIL == 0:
        print("Workbook and CSVs are in sync.")
    else:
        print("MISMATCH detected — workbook and CSVs are out of sync.")
    print("=" * 50)

    return FAIL == 0


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
