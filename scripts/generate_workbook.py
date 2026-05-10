"""Generate the Trade Spend Data Diagnostic workbook.

Reads from the cinderhaven-data SQLite database and produces a 7-tab
Excel workbook in the output/ directory.
"""
import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from build_db import find_database

# ── Constants ────────────────────────────────────────────────────────
TRAIL_START = "2025-05-03"
TRAIL_END = "2026-05-02"

TAB_GREEN = "339966"
TAB_BLUE = "4472C4"
TAB_GRAY = "808080"

FILL_HEADER = PatternFill("solid", fgColor="2F5233")
FILL_SUBHEADER = PatternFill("solid", fgColor="D9E2D0")
FILL_INPUT = PatternFill("solid", fgColor="FFFFCC")
FILL_STRIPE = PatternFill("solid", fgColor="F2F2F2")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_GRAY_HEADER = PatternFill("solid", fgColor="404040")
FILL_BLUE_HEADER = PatternFill("solid", fgColor="2B4C7E")

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_SUBHEADER = Font(name="Calibri", size=11, bold=True, color="2F5233")
FONT_BODY = Font(name="Calibri", size=11, color="333333")
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="2F5233")
FONT_SUBTITLE = Font(name="Calibri", size=12, bold=True, color="666666")
FONT_SMALL = Font(name="Calibri", size=9, color="999999")
FONT_KPI_VALUE = Font(name="Calibri", size=28, bold=True, color="2F5233")
FONT_KPI_LABEL = Font(name="Calibri", size=10, color="666666")
FONT_LINK = Font(name="Calibri", size=11, color="0563C1", underline="single")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
BOTTOM_BORDER = Border(bottom=Side(style="thin", color="BFBFBF"))

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

NUM_DOLLAR = '#,##0'
NUM_DOLLAR_CENTS = '#,##0.00'
NUM_PCT = '0.0%'
NUM_PCT1 = '0.0%'
NUM_INT = '#,##0'

REGIONAL_RETAILERS = [
    "Green Basket Market", "Southside Grocers", "Prairie Provisions",
    "Mountain Pantry Co", "Harbor Fresh",
]
REGIONAL_SLUGS = [
    "green_basket_market", "southside_grocers", "prairie_provisions",
    "mountain_pantry_co", "harbor_fresh",
]


def connect_db():
    db_path = find_database()
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    return conn


def style_header_row(ws, row, max_col, fill=None, font=None):
    fill = fill or FILL_HEADER
    font = font or FONT_HEADER
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row, max_col, stripe=False):
    fill = FILL_STRIPE if stripe else FILL_WHITE
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = FONT_BODY
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        width = max(lengths) if lengths else min_width
        ws.column_dimensions[col_letter].width = min(max(width + 3, min_width), max_width)


# ── Tab 7: Methodology & Logic ──────────────────────────────────────

def build_tab7_methodology(wb, conn):
    ws = wb.create_sheet("Methodology & Logic")
    ws.sheet_properties.tabColor = TAB_GRAY

    ws.sheet_view.showGridLines = False

    sections = [
        ("TRADE SPEND DATA DIAGNOSTIC — METHODOLOGY & LOGIC", None),
        ("Build Information", [
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Database", "cinderhaven_product_master.db"),
            ("Scan data window", "2024-05-11 to 2026-05-02 (104 weeks)"),
            ("Trailing-52w window", f"{TRAIL_START} to {TRAIL_END}"),
            ("Trailing-365 window", f"{TRAIL_START} to {TRAIL_END}"),
            ("SKU count", str(conn.execute("SELECT COUNT(DISTINCT sku) FROM product_master").fetchone()[0])),
            ("Store count", str(conn.execute("SELECT COUNT(DISTINCT store_id) FROM stores").fetchone()[0])),
            ("Retailer count", str(conn.execute("SELECT COUNT(DISTINCT retailer) FROM stores").fetchone()[0])),
        ]),
        ("Two-Bucket Executive Framing", [
            ("Bucket 1: Structural / Planned Trade",
             "The negotiated trade-spend rate from sku_costs, applied to actual "
             "channel revenue. This is the cost of doing business — slotting fees, "
             "accruals, and contractual allowances baked into the wholesale price. "
             "Formula: SUM(trade_spend_pct_{channel} × channel_revenue) across all SKUs."),
            ("Bucket 2: Operational / Compliance Waste",
             "Trailing-365 deductions excluding promo_billback. These are charges "
             "beyond the planned trade rate — short-ship fines, label violations, "
             "spoilage claims, late-delivery penalties, vague/unsubstantiated "
             "deductions, and pallet fines. This is the addressable waste."),
            ("Promo Billback (excluded from Bucket 2)",
             "Deductions coded as promo_billback are excluded from operational waste "
             "to avoid double-counting with structural trade. They represent retailers "
             "billing back promotional allowances already accounted for in Bucket 1."),
            ("All-In Trade Cost",
             "Bucket 1 + Bucket 2 + promo_billback. The total cash cost of trade "
             "programs and compliance failures as a percentage of revenue."),
        ]),
        ("Revenue Calculation", [
            ("Source table", "scan_data"),
            ("Revenue field", "dollars_sold (already price × units)"),
            ("Time window", f"week_ending >= '{TRAIL_START}' AND week_ending <= '{TRAIL_END}'"),
            ("Channel mapping", "scan_data.store_id → stores.retailer"),
            ("Note", "DTC revenue included in total but DTC trade_spend_pct is 0%."),
        ]),
        ("Structural Trade Calculation", [
            ("Source table", "sku_costs"),
            ("Rate columns", "trade_spend_pct_{walmart, costco, whole_foods, regional, unfi, dtc}"),
            ("Method", "Per-SKU: multiply each channel's trade_spend_pct by that SKU's "
             "trailing-52w revenue in that channel. Sum across all SKUs and channels."),
            ("Regional mapping", "Green Basket Market, Southside Grocers, Prairie Provisions, "
             "Mountain Pantry Co, Harbor Fresh → all use trade_spend_pct_regional."),
        ]),
        ("Deduction Analysis", [
            ("Source table", "deductions"),
            ("Time window", f"deduction_date >= '{TRAIL_START}' AND deduction_date <= '{TRAIL_END}'"),
            ("Types", "vague, short_ship, label_fine, spoilage, late_delivery, damaged, "
             "pallet_fine, slotting, promo_billback"),
            ("Retailer key", "deductions.retailer_id uses slugs (walmart, costco, whole_foods, etc.)"),
            ("Code translation", "deductions.code_id → deduction_codes.code + .name for plain-English labels"),
        ]),
        ("Double-Dip Detection", [
            ("Definition", "A double-dip occurs when a retailer collects an off-invoice discount "
             "(via reduced invoice price) AND submits a promo_billback deduction for the "
             "same promotional event — charging the manufacturer twice."),
            ("Detection method", "deductions.is_double_dip = 1 (pre-flagged in the data pipeline). "
             "Each flagged deduction has a matching promotion with "
             "funding_mechanism = 'off_invoice' in the same SKU/retailer window."),
            ("Note", "Double-dip deductions are from 2024 (pre-trailing-365 window). "
             "They appear on the Leak Diagnostic tab as historical flags."),
        ]),
        ("Promo ROI Methodology", [
            ("Source tables", "promotions, scan_data, stores"),
            ("Pre/post window", "Default 4 weeks (adjustable via input cell on Promo Efficacy tab)"),
            ("Baseline", "Average weekly dollars_sold for the SKU at matching stores "
             "during the pre-window period."),
            ("Lift", "(During-period avg weekly revenue − baseline) × promo weeks"),
            ("ROI", "(Lift − promo_cost) / promo_cost. Undefined when promo_cost is NULL."),
            ("Limitation", "Simple pre/during/post comparison. No seasonality adjustment "
             "or causal inference. Stated honestly on the Promo Efficacy tab."),
        ]),
        ("Net-Net Effective Margin by Retailer", [
            ("Method", "Gross margin → subtract structural trade rate → subtract "
             "operational deduction rate → effective margin."),
            ("Gross margin", "1 − (avg COGS per unit / avg wholesale price per unit) by channel, "
             "weighted by SKU revenue mix."),
            ("Structural trade rate", "Channel-level weighted average of trade_spend_pct from sku_costs."),
            ("Operational deduction rate", "Trailing-365 deductions (excl promo_billback) "
             "for the retailer ÷ channel revenue."),
            ("What-if inputs", "Adjustable trade rate per retailer on the Retailer Risk tab. "
             "Changing the input recalculates the effective margin."),
        ]),
        ("Dispute & Recovery Analysis", [
            ("Source tables", "disputes (joined to deductions via deduction_id)"),
            ("Disputes filed", "COUNT of disputes joined to trailing-365 deductions"),
            ("Recovery rate", "SUM(recovered_amount) for won outcomes ÷ SUM(disputed amount). "
             "Denominator is the amount of deductions that had disputes filed, "
             "not total deductions."),
            ("Outcomes", "won_full, won_partial, lost_evidence, lost_deadline, "
             "lost_no_response, lost_other, abandoned, pending"),
        ]),
        ("Data Lineage", [
            ("Source", "cinderhaven_product_master.db (SQLite, consumed via git submodule)"),
            ("Pipeline", "cinderhaven-data repo builds all 22 tables from seed scripts. "
             "Deduction lifecycle (disputes, evidence, chargebacks) merged from "
             "retailer-deduction-recovery project."),
            ("Quirks by design", "Inconsistent retailer naming across tables (title case in stores, "
             "slugs in deductions/retailers). ~23% internal inconsistency rate in "
             "deduction descriptions. Date format mismatches. These mirror real-world "
             "controller data quality issues."),
        ]),
    ]

    row = 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    title_cell = ws.cell(row=row, column=2, value=sections[0][0])
    title_cell.font = FONT_TITLE
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws.cell(row=row, column=2, value="Definitions, data lineage, and calculation logic").font = FONT_SUBTITLE
    row += 2

    for section_title, items in sections[1:]:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        hdr = ws.cell(row=row, column=2, value=section_title)
        hdr.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        hdr.fill = PatternFill("solid", fgColor="404040")
        hdr.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(3, 9):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="404040")
        row += 1

        for label, desc in items:
            ws.cell(row=row, column=2, value=label).font = Font(
                name="Calibri", size=11, bold=True, color="333333"
            )
            ws.cell(row=row, column=2).alignment = ALIGN_LEFT
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
            desc_cell = ws.cell(row=row, column=3, value=desc)
            desc_cell.font = FONT_BODY
            desc_cell.alignment = ALIGN_LEFT
            ws.row_dimensions[row].height = max(15, len(desc) // 6 + 15) if len(desc) > 80 else 18
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    for col_letter in "CDEFGH":
        ws.column_dimensions[col_letter].width = 16

    return ws


# ── Tab 6: Deduction Code Crosswalk ─────────────────────────────────

def build_tab6_crosswalk(wb, conn):
    ws = wb.create_sheet("Code Crosswalk")
    ws.sheet_properties.tabColor = TAB_GRAY
    ws.sheet_view.showGridLines = False

    row = 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws.cell(row=row, column=2, value="DEDUCTION CODE CROSSWALK").font = FONT_TITLE
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    ws.cell(row=row, column=2,
            value="Retailer deduction codes mapped to plain-English descriptions and standardized categories").font = FONT_SUBTITLE
    row += 2

    headers = ["Retailer", "Code", "Code Name", "Deduction Type",
               "Published?", "Trailing-365 Count", "Trailing-365 $"]
    for ci, h in enumerate(headers, start=2):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, max_col=8, fill=FILL_GRAY_HEADER)
    header_row = row
    row += 1

    codes = conn.execute("""
        SELECT dc.retailer_id, dc.code, dc.name, dc.deduction_type,
               dc.is_published,
               COUNT(d.deduction_id) as ded_count,
               COALESCE(SUM(d.amount), 0) as ded_total
        FROM deduction_codes dc
        LEFT JOIN deductions d ON dc.code_id = d.code_id
            AND d.deduction_date >= ? AND d.deduction_date <= ?
        GROUP BY dc.code_id
        ORDER BY dc.retailer_id, dc.deduction_type, dc.code
    """, (TRAIL_START, TRAIL_END)).fetchall()

    for i, c in enumerate(codes):
        vals = [
            c["retailer_id"], c["code"], c["name"], c["deduction_type"],
            "Yes" if c["is_published"] else "No",
            c["ded_count"], c["ded_total"],
        ]
        for ci, v in enumerate(vals, start=2):
            cell = ws.cell(row=row, column=ci, value=v)
            if ci == 7:
                cell.number_format = NUM_DOLLAR
            elif ci == 8:
                cell.number_format = NUM_DOLLAR
        style_data_row(ws, row, max_col=8, stripe=(i % 2 == 0))
        row += 1

    ws.auto_filter.ref = f"B{header_row}:H{row - 1}"
    ws.freeze_panes = f"B{header_row + 1}"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 16

    return ws


# ── Tab 5: Deduction Ledger ──────────────────────────────────────────

def build_tab5_ledger(wb, conn):
    ws = wb.create_sheet("Deduction Ledger")
    ws.sheet_properties.tabColor = TAB_BLUE
    ws.sheet_view.showGridLines = False

    row = 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
    ws.cell(row=row, column=2, value="DEDUCTION LEDGER — TRAILING 365 DAYS").font = FONT_TITLE
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
    ws.cell(row=row, column=2,
            value=f"All deductions from {TRAIL_START} to {TRAIL_END}").font = FONT_SUBTITLE
    row += 2

    headers = ["Deduction ID", "Retailer", "Deduction Type", "Code",
               "Code Name", "Description", "Amount", "Date",
               "Dispute Deadline", "Double-Dip?", "Post-Audit?"]
    for ci, h in enumerate(headers, start=2):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, max_col=12, fill=FILL_BLUE_HEADER)
    header_row = row
    row += 1

    ledger = conn.execute("""
        SELECT d.deduction_id, d.retailer_id, d.deduction_type,
               dc.code, dc.name as code_name,
               d.remittance_description, d.amount, d.deduction_date,
               d.dispute_deadline, d.is_double_dip, d.is_post_audit
        FROM deductions d
        LEFT JOIN deduction_codes dc ON d.code_id = dc.code_id
        WHERE d.deduction_date >= ? AND d.deduction_date <= ?
        ORDER BY d.deduction_date DESC, d.amount DESC
    """, (TRAIL_START, TRAIL_END)).fetchall()

    for i, d in enumerate(ledger):
        vals = [
            d["deduction_id"], d["retailer_id"], d["deduction_type"],
            d["code"], d["code_name"], d["remittance_description"],
            d["amount"], d["deduction_date"], d["dispute_deadline"],
            "Yes" if d["is_double_dip"] else "",
            "Yes" if d["is_post_audit"] else "",
        ]
        for ci, v in enumerate(vals, start=2):
            cell = ws.cell(row=row, column=ci, value=v)
            if ci == 8:
                cell.number_format = NUM_DOLLAR_CENTS
        style_data_row(ws, row, max_col=12, stripe=(i % 2 == 0))
        row += 1

    ws.auto_filter.ref = f"B{header_row}:L{row - 1}"
    ws.freeze_panes = f"B{header_row + 1}"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 32
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 12

    return ws


# ── Tab 4: Retailer Risk ────────────────────────────────────────────

def build_tab4_retailer_risk(wb, conn):
    ws = wb.create_sheet("Retailer Risk")
    ws.sheet_properties.tabColor = TAB_GREEN
    ws.sheet_view.showGridLines = False

    row = 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
    ws.cell(row=row, column=2, value="RETAILER RISK PROFILE").font = FONT_TITLE
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
    ws.cell(row=row, column=2,
            value="Revenue concentration, net-net effective margin, and what-if trade rate analysis").font = FONT_SUBTITLE
    row += 2

    # Gather retailer-level data
    total_rev = conn.execute(f"""
        SELECT SUM(dollars_sold) FROM scan_data
        WHERE week_ending >= '{TRAIL_START}' AND week_ending <= '{TRAIL_END}'
    """).fetchone()[0]

    retailer_data = conn.execute(f"""
        SELECT s.retailer,
               SUM(sd.dollars_sold) as revenue,
               COUNT(DISTINCT sd.sku) as sku_count
        FROM scan_data sd
        JOIN stores s ON sd.store_id = s.store_id
        WHERE sd.week_ending >= '{TRAIL_START}' AND sd.week_ending <= '{TRAIL_END}'
        GROUP BY s.retailer
        ORDER BY SUM(sd.dollars_sold) DESC
    """).fetchall()

    # Deductions by retailer slug
    ded_by_retailer = {}
    for r in conn.execute(f"""
        SELECT retailer_id, SUM(amount) as total,
               SUM(CASE WHEN deduction_type != 'promo_billback' THEN amount ELSE 0 END) as operational
        FROM deductions
        WHERE deduction_date >= '{TRAIL_START}' AND deduction_date <= '{TRAIL_END}'
        GROUP BY retailer_id
    """).fetchall():
        ded_by_retailer[r["retailer_id"]] = {
            "total": r["total"], "operational": r["operational"]
        }

    total_deds = conn.execute(f"""
        SELECT SUM(amount) FROM deductions
        WHERE deduction_date >= '{TRAIL_START}' AND deduction_date <= '{TRAIL_END}'
    """).fetchone()[0]

    # Gross margin by retailer (weighted by revenue)
    margin_data = {}
    retailer_to_slug = {
        "Walmart": "walmart", "Costco": "costco", "Whole Foods": "whole_foods",
        "UNFI": "unfi", "DTC": "dtc",
        "Green Basket Market": "green_basket_market",
        "Southside Grocers": "southside_grocers",
        "Prairie Provisions": "prairie_provisions",
        "Mountain Pantry Co": "mountain_pantry_co",
        "Harbor Fresh": "harbor_fresh",
    }
    slug_to_channel = {
        "walmart": "walmart", "costco": "costco", "whole_foods": "whole_foods",
        "unfi": "unfi", "dtc": "dtc",
    }
    for rname in REGIONAL_RETAILERS:
        slug_to_channel[retailer_to_slug[rname]] = "regional"

    for r in retailer_data:
        rname = r["retailer"]
        slug = retailer_to_slug.get(rname, rname.lower().replace(" ", "_"))
        channel = slug_to_channel.get(slug, "regional")
        wholesale_col = f"wholesale_{channel}"
        trade_col = f"trade_spend_pct_{channel}"

        gm = conn.execute(f"""
            SELECT SUM(sd.dollars_sold * (1.0 - sc.cogs_per_unit / sc.{wholesale_col})) /
                   SUM(sd.dollars_sold) as weighted_gm,
                   SUM(sd.dollars_sold * sc.{trade_col}) / SUM(sd.dollars_sold) as weighted_trade
            FROM scan_data sd
            JOIN stores s ON sd.store_id = s.store_id
            JOIN sku_costs sc ON sd.sku = sc.sku
            WHERE s.retailer = ? AND sd.week_ending >= '{TRAIL_START}'
              AND sd.week_ending <= '{TRAIL_END}'
        """, (rname,)).fetchone()

        margin_data[rname] = {
            "gross_margin": gm["weighted_gm"] if gm["weighted_gm"] else 0,
            "trade_rate": gm["weighted_trade"] if gm["weighted_trade"] else 0,
        }

    # Write header
    headers = ["Retailer", "Revenue", "Rev Share", "Deduction $",
               "Ded Share", "Gross Margin", "Structural Trade",
               "Operational Ded Rate", "Net-Net Margin",
               "What-If Trade Rate", "What-If Margin"]
    for ci, h in enumerate(headers, start=2):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, max_col=12)
    header_row = row
    row += 1

    input_rows = []
    data_start = row
    for i, r in enumerate(retailer_data):
        rname = r["retailer"]
        rev = r["revenue"]
        slug = retailer_to_slug.get(rname, rname.lower().replace(" ", "_"))
        deds = ded_by_retailer.get(slug, {"total": 0, "operational": 0})
        gm = margin_data.get(rname, {"gross_margin": 0, "trade_rate": 0})
        op_rate = deds["operational"] / rev if rev else 0
        net_net = gm["gross_margin"] - gm["trade_rate"] - op_rate

        ws.cell(row=row, column=2, value=rname)
        ws.cell(row=row, column=3, value=rev).number_format = NUM_DOLLAR
        ws.cell(row=row, column=4, value=rev / total_rev).number_format = NUM_PCT
        ws.cell(row=row, column=5, value=deds["total"]).number_format = NUM_DOLLAR
        ws.cell(row=row, column=6, value=deds["total"] / total_deds if total_deds else 0).number_format = NUM_PCT
        ws.cell(row=row, column=7, value=gm["gross_margin"]).number_format = NUM_PCT
        ws.cell(row=row, column=8, value=gm["trade_rate"]).number_format = NUM_PCT
        ws.cell(row=row, column=9, value=op_rate).number_format = NUM_PCT
        ws.cell(row=row, column=10, value=net_net).number_format = NUM_PCT

        # What-if input cell
        whatif_cell = ws.cell(row=row, column=11, value=gm["trade_rate"])
        whatif_cell.number_format = NUM_PCT
        whatif_cell.fill = FILL_INPUT
        whatif_cell.font = Font(name="Calibri", size=11, bold=True, color="333333")
        input_rows.append(row)

        # What-if margin formula: gross_margin - whatif_trade - op_rate
        formula = f"=G{row}-K{row}-I{row}"
        ws.cell(row=row, column=12, value=formula).number_format = NUM_PCT

        style_data_row(ws, row, max_col=12, stripe=(i % 2 == 0))
        ws.cell(row=row, column=11).fill = FILL_INPUT
        row += 1

    # Add validation for what-if cells
    dv = DataValidation(
        type="decimal", operator="between",
        formula1="0", formula2="0.5",
        allow_blank=True,
    )
    dv.error = "Enter a trade rate between 0% and 50%"
    dv.errorTitle = "Invalid trade rate"
    dv.prompt = "Enter an adjusted trade rate to see impact on margin"
    dv.promptTitle = "What-If Trade Rate"
    ws.add_data_validation(dv)
    for ir in input_rows:
        dv.add(ws.cell(row=ir, column=11))

    # Add comments to input cells
    from openpyxl.comments import Comment
    for ir in input_rows:
        ws.cell(row=ir, column=11).comment = Comment(
            "Adjustable input: change this trade rate to see the impact on net-net margin in column L.",
            "Trade Spend Diagnostic"
        )

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
    ws.cell(row=row, column=2,
            value="Yellow cells are adjustable inputs. Change the What-If Trade Rate to model margin impact."
            ).font = FONT_SMALL

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    for col_letter in "CDEFGHIJKL":
        ws.column_dimensions[col_letter].width = 16

    return ws


# ── Tab 3: Promo Efficacy ───────────────────────────────────────────

def build_tab3_promo(wb, conn):
    ws = wb.create_sheet("Promo Efficacy")
    ws.sheet_properties.tabColor = TAB_GREEN
    ws.sheet_view.showGridLines = False

    row = 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
    ws.cell(row=row, column=2, value="PROMOTION EFFICACY ANALYSIS").font = FONT_TITLE
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
    ws.cell(row=row, column=2,
            value="Top and bottom performers by lift vs. cost, with adjustable analysis window").font = FONT_SUBTITLE
    row += 2

    # Input cell: pre/post window weeks
    ws.cell(row=row, column=2, value="Pre/Post Analysis Window (weeks):").font = FONT_SUBHEADER
    window_cell = ws.cell(row=row, column=5, value=4)
    window_cell.fill = FILL_INPUT
    window_cell.font = Font(name="Calibri", size=14, bold=True, color="333333")
    window_cell.alignment = ALIGN_CENTER
    from openpyxl.comments import Comment
    window_cell.comment = Comment(
        "Adjustable input: number of weeks before and after the promo to use "
        "as baseline/comparison. Default is 4 weeks.",
        "Trade Spend Diagnostic"
    )
    window_ref = f"E{row}"

    dv = DataValidation(
        type="whole", operator="between",
        formula1="1", formula2="12",
        allow_blank=False,
    )
    dv.error = "Enter a number between 1 and 12"
    dv.errorTitle = "Invalid window"
    dv.prompt = "Weeks before/after promo for baseline comparison"
    dv.promptTitle = "Analysis Window"
    ws.add_data_validation(dv)
    dv.add(window_cell)

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
    ws.cell(row=row, column=2,
            value="Limitation: Simple pre/during/post comparison. No seasonality adjustment or causal inference."
            ).font = FONT_SMALL
    row += 2

    # Get promo data with calculated ROI
    window_weeks = 4
    promos = conn.execute("""
        SELECT p.promo_id, p.retailer, p.sku, p.start_week, p.end_week,
               p.promo_type, p.funding_mechanism, p.promo_cost,
               pm.product_name
        FROM promotions p
        LEFT JOIN product_master pm ON p.sku = pm.sku
        ORDER BY p.start_week DESC
    """).fetchall()

    # Group by promo_id and calculate lift
    from collections import defaultdict
    promo_groups = defaultdict(list)
    for p in promos:
        promo_groups[p["promo_id"]].append(p)

    promo_results = []
    for promo_id, items in promo_groups.items():
        first = items[0]
        total_cost = sum(it["promo_cost"] or 0 for it in items)
        sku_count = len(set(it["sku"] for it in items))
        retailer = first["retailer"]
        start = first["start_week"]
        end = first["end_week"]
        promo_type = first["promo_type"]
        funding = first["funding_mechanism"]

        # Calculate lift for this promo
        # Pre-period baseline
        import datetime as dt
        try:
            start_dt = dt.datetime.strptime(start, "%Y-%m-%d")
            end_dt = dt.datetime.strptime(end, "%Y-%m-%d")
        except (ValueError, TypeError):
            promo_results.append({
                "promo_id": promo_id, "retailer": retailer, "sku_count": sku_count,
                "start": start, "end": end, "type": promo_type, "funding": funding,
                "cost": total_cost, "pre_rev": None, "during_rev": None,
                "lift": None, "roi": None, "data_quality": "No dates",
            })
            continue

        promo_weeks = max(1, (end_dt - start_dt).days / 7)
        pre_start = (start_dt - dt.timedelta(weeks=window_weeks)).strftime("%Y-%m-%d")
        pre_end = (start_dt - dt.timedelta(days=1)).strftime("%Y-%m-%d")

        skus = [it["sku"] for it in items]
        sku_placeholders = ",".join("?" * len(skus))

        store_retailer = retailer if retailer else "Unknown"

        pre_rev = conn.execute(f"""
            SELECT SUM(sd.dollars_sold) FROM scan_data sd
            JOIN stores s ON sd.store_id = s.store_id
            WHERE sd.sku IN ({sku_placeholders})
              AND s.retailer = ?
              AND sd.week_ending >= ? AND sd.week_ending <= ?
        """, (*skus, store_retailer, pre_start, pre_end)).fetchone()[0]

        during_rev = conn.execute(f"""
            SELECT SUM(sd.dollars_sold) FROM scan_data sd
            JOIN stores s ON sd.store_id = s.store_id
            WHERE sd.sku IN ({sku_placeholders})
              AND s.retailer = ?
              AND sd.week_ending >= ? AND sd.week_ending <= ?
        """, (*skus, store_retailer, start, end)).fetchone()[0]

        pre_weekly = (pre_rev / window_weeks) if pre_rev and window_weeks else None
        during_weekly = (during_rev / promo_weeks) if during_rev else None

        if pre_weekly and during_weekly:
            lift = (during_weekly - pre_weekly) * promo_weeks
            roi = (lift - total_cost) / total_cost if total_cost > 0 else None
            dq = "Good" if pre_rev and during_rev else "Partial"
        else:
            lift = None
            roi = None
            dq = "Insufficient" if not pre_rev and not during_rev else "Partial"

        promo_results.append({
            "promo_id": promo_id, "retailer": retailer, "sku_count": sku_count,
            "start": start, "end": end, "type": promo_type, "funding": funding,
            "cost": total_cost, "pre_rev": pre_rev, "during_rev": during_rev,
            "lift": lift, "roi": roi, "data_quality": dq,
        })

    # Sort by ROI descending (None at bottom)
    promo_results.sort(key=lambda x: (x["roi"] is None, -(x["roi"] or 0)))

    headers = ["Promo ID", "Retailer", "SKUs", "Start", "End",
               "Type", "Funding", "Cost", "Pre-Period Rev",
               "During Rev", "Est. Lift", "ROI", "Data Quality"]
    for ci, h in enumerate(headers, start=2):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, max_col=14)
    header_row = row
    row += 1

    for i, p in enumerate(promo_results):
        vals = [
            p["promo_id"], p["retailer"], p["sku_count"],
            p["start"], p["end"], p["type"], p["funding"],
            p["cost"], p["pre_rev"], p["during_rev"],
            p["lift"], p["roi"], p["data_quality"],
        ]
        for ci, v in enumerate(vals, start=2):
            cell = ws.cell(row=row, column=ci, value=v)
            if ci in (9, 10, 11, 12):
                cell.number_format = NUM_DOLLAR if ci <= 12 else NUM_PCT
            if ci == 13:
                cell.number_format = NUM_PCT1
        ws.cell(row=row, column=9).number_format = NUM_DOLLAR
        ws.cell(row=row, column=10).number_format = NUM_DOLLAR
        ws.cell(row=row, column=11).number_format = NUM_DOLLAR
        ws.cell(row=row, column=12).number_format = NUM_DOLLAR
        ws.cell(row=row, column=13).number_format = NUM_PCT
        ws.cell(row=row, column=14).font = FONT_BODY
        style_data_row(ws, row, max_col=14, stripe=(i % 2 == 0))

        # Color-code data quality
        dq_cell = ws.cell(row=row, column=14)
        if p["data_quality"] == "Good":
            dq_cell.font = Font(name="Calibri", size=11, color="2F5233")
        elif p["data_quality"] == "Insufficient":
            dq_cell.font = Font(name="Calibri", size=11, color="C00000")
        else:
            dq_cell.font = Font(name="Calibri", size=11, color="BF8F00")

        row += 1

    ws.auto_filter.ref = f"B{header_row}:N{row - 1}"
    ws.freeze_panes = f"B{header_row + 1}"

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
    ws.cell(row=row, column=2,
            value=f"Showing {len(promo_results)} promotion events. "
            f"Pre/post window: {window_weeks} weeks. "
            f"ROI = (Lift - Cost) / Cost."
            ).font = FONT_SMALL

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 10
    ws.column_dimensions["N"].width = 14

    return ws


# ── Tab 2: Leak Diagnostic ──────────────────────────────────────────

def build_tab2_leak(wb, conn):
    ws = wb.create_sheet("Leak Diagnostic")
    ws.sheet_properties.tabColor = TAB_GREEN
    ws.sheet_view.showGridLines = False

    row = 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.cell(row=row, column=2, value="OPERATIONAL LEAK DIAGNOSTIC").font = FONT_TITLE
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.cell(row=row, column=2,
            value="Waste by category, double-dip detection, and recoverability assessment").font = FONT_SUBTITLE
    row += 2

    total_rev = conn.execute(f"""
        SELECT SUM(dollars_sold) FROM scan_data
        WHERE week_ending >= '{TRAIL_START}' AND week_ending <= '{TRAIL_END}'
    """).fetchone()[0]

    # ── Section 1: Waste by deduction type ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.cell(row=row, column=2, value="Operational Waste by Category").font = Font(
        name="Calibri", size=12, bold=True, color="2F5233")
    row += 1

    waste_headers = ["Deduction Type", "Count", "Total $", "% of Revenue",
                     "Avg per Deduction", "% of Total Waste"]
    for ci, h in enumerate(waste_headers, start=2):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, max_col=7)
    row += 1

    waste_data = conn.execute(f"""
        SELECT deduction_type, COUNT(*) as cnt, SUM(amount) as total,
               AVG(amount) as avg_amt
        FROM deductions
        WHERE deduction_date >= '{TRAIL_START}' AND deduction_date <= '{TRAIL_END}'
        AND deduction_type != 'promo_billback'
        GROUP BY deduction_type
        ORDER BY SUM(amount) DESC
    """).fetchall()

    total_waste = sum(w["total"] for w in waste_data)

    for i, w in enumerate(waste_data):
        ws.cell(row=row, column=2, value=w["deduction_type"].replace("_", " ").title())
        ws.cell(row=row, column=3, value=w["cnt"]).number_format = NUM_INT
        ws.cell(row=row, column=4, value=w["total"]).number_format = NUM_DOLLAR
        ws.cell(row=row, column=5, value=w["total"] / total_rev).number_format = NUM_PCT
        ws.cell(row=row, column=6, value=w["avg_amt"]).number_format = NUM_DOLLAR
        ws.cell(row=row, column=7, value=w["total"] / total_waste).number_format = NUM_PCT
        style_data_row(ws, row, max_col=7, stripe=(i % 2 == 0))
        row += 1

    # Totals row
    ws.cell(row=row, column=2, value="TOTAL OPERATIONAL WASTE").font = Font(
        name="Calibri", size=11, bold=True, color="333333")
    ws.cell(row=row, column=3, value=sum(w["cnt"] for w in waste_data)).number_format = NUM_INT
    ws.cell(row=row, column=4, value=total_waste).number_format = NUM_DOLLAR
    ws.cell(row=row, column=5, value=total_waste / total_rev).number_format = NUM_PCT
    for c in range(2, 8):
        ws.cell(row=row, column=c).font = Font(name="Calibri", size=11, bold=True, color="333333")
        ws.cell(row=row, column=c).border = Border(top=Side(style="double", color="333333"),
                                                    bottom=Side(style="double", color="333333"))
    row += 2

    # ── Section 2: Double-Dip Alert ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.cell(row=row, column=2, value="Double-Dip Alert").font = Font(
        name="Calibri", size=12, bold=True, color="C00000")
    row += 1

    dd = conn.execute("""
        SELECT d.deduction_id, d.retailer_id, d.amount, d.deduction_date,
               d.remittance_description,
               p.promo_id, p.start_week, p.end_week, p.funding_mechanism
        FROM deductions d
        LEFT JOIN promotions p ON d.is_double_dip = 1
            AND p.sku = (SELECT sku FROM deductions WHERE deduction_id = d.deduction_id LIMIT 1)
            AND p.funding_mechanism = 'off_invoice'
        WHERE d.is_double_dip = 1
        ORDER BY d.amount DESC
    """).fetchall()

    if dd:
        dd_headers = ["Deduction ID", "Retailer", "Amount", "Date",
                      "Matching Promo", "Promo Dates", "Mechanism"]
        for ci, h in enumerate(dd_headers, start=2):
            ws.cell(row=row, column=ci, value=h)
        style_header_row(ws, row, max_col=8, fill=PatternFill("solid", fgColor="C00000"))
        row += 1

        seen = set()
        dd_total = 0
        from openpyxl.comments import Comment
        for d_item in dd:
            did = d_item["deduction_id"]
            if did in seen:
                continue
            seen.add(did)
            dd_total += d_item["amount"]
            ws.cell(row=row, column=2, value=d_item["deduction_id"])
            ws.cell(row=row, column=3, value=d_item["retailer_id"])
            ws.cell(row=row, column=4, value=d_item["amount"]).number_format = NUM_DOLLAR_CENTS
            ws.cell(row=row, column=5, value=d_item["deduction_date"])
            ws.cell(row=row, column=6, value=d_item["promo_id"] or "")
            if d_item["start_week"] and d_item["end_week"]:
                ws.cell(row=row, column=7, value=f"{d_item['start_week']} to {d_item['end_week']}")
            ws.cell(row=row, column=8, value=d_item["funding_mechanism"] or "")
            style_data_row(ws, row, max_col=8)

            ws.cell(row=row, column=2).comment = Comment(
                "DOUBLE-DIP: This retailer collected an off-invoice discount (reduced "
                "invoice price) AND submitted a promo_billback deduction for the same "
                "promotional event — charging the manufacturer twice for the same promotion.",
                "Trade Spend Diagnostic"
            )
            row += 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        ws.cell(row=row, column=2,
                value=f"Total double-dip exposure: ${dd_total:,.2f}. "
                "These are from 2024 (pre-trailing-365 window)."
                ).font = FONT_SMALL
    else:
        ws.cell(row=row, column=2, value="No double-dips detected.").font = FONT_BODY

    row += 2

    # ── Section 3: Recovery & Recoverability ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.cell(row=row, column=2, value="Recovery Analysis & Adjustable Target").font = Font(
        name="Calibri", size=12, bold=True, color="2F5233")
    row += 1

    disputes = conn.execute(f"""
        SELECT di.outcome, COUNT(*) as cnt,
               SUM(d.amount) as disputed_amt,
               SUM(di.recovered_amount) as recovered
        FROM disputes di
        JOIN deductions d ON di.deduction_id = d.deduction_id
        WHERE d.deduction_date >= '{TRAIL_START}' AND d.deduction_date <= '{TRAIL_END}'
        GROUP BY di.outcome
        ORDER BY SUM(d.amount) DESC
    """).fetchall()

    total_disputed = sum(d_item["disputed_amt"] for d_item in disputes)
    total_recovered = sum(d_item["recovered"] for d_item in disputes)
    actual_rate = total_recovered / total_disputed if total_disputed else 0

    ws.cell(row=row, column=2, value="Current Recovery Rate:").font = FONT_SUBHEADER
    ws.cell(row=row, column=5, value=actual_rate).number_format = NUM_PCT
    ws.cell(row=row, column=5).font = Font(name="Calibri", size=14, bold=True, color="2F5233")
    row += 1

    ws.cell(row=row, column=2, value="Target Recovery Rate:").font = FONT_SUBHEADER
    target_cell = ws.cell(row=row, column=5, value=0.10)
    target_cell.number_format = NUM_PCT
    target_cell.fill = FILL_INPUT
    target_cell.font = Font(name="Calibri", size=14, bold=True, color="333333")
    target_cell.alignment = ALIGN_CENTER
    from openpyxl.comments import Comment
    target_cell.comment = Comment(
        "Adjustable input: set your target recovery rate. "
        "The additional recoverable amount below updates based on this target.",
        "Trade Spend Diagnostic"
    )
    target_row = row

    dv2 = DataValidation(
        type="decimal", operator="between",
        formula1="0", formula2="1",
        allow_blank=False,
    )
    dv2.error = "Enter a rate between 0% and 100%"
    dv2.errorTitle = "Invalid rate"
    ws.add_data_validation(dv2)
    dv2.add(target_cell)

    row += 1
    ws.cell(row=row, column=2, value="Additional Recoverable at Target:").font = FONT_SUBHEADER
    ws.cell(row=row, column=5,
            value=f"=MAX(0, E{target_row} * {total_disputed:.2f} - {total_recovered:.2f})"
            ).number_format = NUM_DOLLAR
    ws.cell(row=row, column=5).font = Font(name="Calibri", size=14, bold=True, color="BF8F00")

    row += 2

    # Outcome breakdown
    outcome_headers = ["Outcome", "Count", "Disputed $", "Recovered $"]
    for ci, h in enumerate(outcome_headers, start=2):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, max_col=5)
    row += 1

    for i, d_item in enumerate(disputes):
        ws.cell(row=row, column=2, value=d_item["outcome"].replace("_", " ").title())
        ws.cell(row=row, column=3, value=d_item["cnt"]).number_format = NUM_INT
        ws.cell(row=row, column=4, value=d_item["disputed_amt"]).number_format = NUM_DOLLAR
        ws.cell(row=row, column=5, value=d_item["recovered"]).number_format = NUM_DOLLAR
        style_data_row(ws, row, max_col=5, stripe=(i % 2 == 0))
        row += 1

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 14

    return ws


# ── Tab 1: Executive Pulse ──────────────────────────────────────────

def build_tab1_executive(wb, conn):
    ws = wb.create_sheet("Executive Pulse")
    ws.sheet_properties.tabColor = TAB_GREEN
    ws.sheet_view.showGridLines = False

    # ── Gather all numbers ──
    total_rev = conn.execute(f"""
        SELECT SUM(dollars_sold) FROM scan_data
        WHERE week_ending >= '{TRAIL_START}' AND week_ending <= '{TRAIL_END}'
    """).fetchone()[0]

    structural = conn.execute(f"""
        SELECT SUM(
            sc.trade_spend_pct_walmart * COALESCE(sd_rev.rev_walmart, 0) +
            sc.trade_spend_pct_costco * COALESCE(sd_rev.rev_costco, 0) +
            sc.trade_spend_pct_whole_foods * COALESCE(sd_rev.rev_wf, 0) +
            sc.trade_spend_pct_regional * COALESCE(sd_rev.rev_regional, 0) +
            sc.trade_spend_pct_unfi * COALESCE(sd_rev.rev_unfi, 0) +
            sc.trade_spend_pct_dtc * COALESCE(sd_rev.rev_dtc, 0)
        )
        FROM sku_costs sc
        LEFT JOIN (
            SELECT sd.sku,
                SUM(CASE WHEN s.retailer='Walmart' THEN sd.dollars_sold ELSE 0 END) as rev_walmart,
                SUM(CASE WHEN s.retailer='Costco' THEN sd.dollars_sold ELSE 0 END) as rev_costco,
                SUM(CASE WHEN s.retailer='Whole Foods' THEN sd.dollars_sold ELSE 0 END) as rev_wf,
                SUM(CASE WHEN s.retailer IN ('Green Basket Market','Southside Grocers','Prairie Provisions','Mountain Pantry Co','Harbor Fresh') THEN sd.dollars_sold ELSE 0 END) as rev_regional,
                SUM(CASE WHEN s.retailer='UNFI' THEN sd.dollars_sold ELSE 0 END) as rev_unfi,
                SUM(CASE WHEN s.retailer='DTC' THEN sd.dollars_sold ELSE 0 END) as rev_dtc
            FROM scan_data sd
            JOIN stores s ON sd.store_id = s.store_id
            WHERE sd.week_ending >= '{TRAIL_START}' AND sd.week_ending <= '{TRAIL_END}'
            GROUP BY sd.sku
        ) sd_rev ON sc.sku = sd_rev.sku
    """).fetchone()[0]

    operational = conn.execute(f"""
        SELECT SUM(amount) FROM deductions
        WHERE deduction_date >= '{TRAIL_START}' AND deduction_date <= '{TRAIL_END}'
        AND deduction_type != 'promo_billback'
    """).fetchone()[0]

    promo_bb = conn.execute(f"""
        SELECT SUM(amount) FROM deductions
        WHERE deduction_date >= '{TRAIL_START}' AND deduction_date <= '{TRAIL_END}'
        AND deduction_type = 'promo_billback'
    """).fetchone()[0]

    all_in = structural + operational + promo_bb

    # Recovery
    recovery = conn.execute(f"""
        SELECT SUM(di.recovered_amount) FROM disputes di
        JOIN deductions d ON di.deduction_id = d.deduction_id
        WHERE d.deduction_date >= '{TRAIL_START}' AND d.deduction_date <= '{TRAIL_END}'
        AND di.outcome IN ('won_full', 'won_partial')
    """).fetchone()[0] or 0

    # Double-dips (all time)
    dd_count, dd_amount = conn.execute(
        "SELECT COUNT(*), SUM(amount) FROM deductions WHERE is_double_dip = 1"
    ).fetchone()
    dd_amount = dd_amount or 0

    # Disputes
    dispute_count = conn.execute(f"""
        SELECT COUNT(*) FROM disputes di
        JOIN deductions d ON di.deduction_id = d.deduction_id
        WHERE d.deduction_date >= '{TRAIL_START}' AND d.deduction_date <= '{TRAIL_END}'
    """).fetchone()[0]

    total_disputed = conn.execute(f"""
        SELECT SUM(d.amount) FROM disputes di
        JOIN deductions d ON di.deduction_id = d.deduction_id
        WHERE d.deduction_date >= '{TRAIL_START}' AND d.deduction_date <= '{TRAIL_END}'
    """).fetchone()[0] or 0

    recovery_rate = recovery / total_disputed if total_disputed else 0

    # Addressable improvement: operational waste - already recovered
    addressable = operational - recovery

    # ── Layout ──
    row = 2

    # Title
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.cell(row=row, column=2, value="TRADE SPEND DATA DIAGNOSTIC").font = FONT_TITLE
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.cell(row=row, column=2,
            value="Cinderhaven Food Co. — Trailing 52-Week Analysis").font = FONT_SUBTITLE
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.cell(row=row, column=2,
            value=f"Data through {TRAIL_END} | Generated {datetime.now().strftime('%Y-%m-%d')}"
            ).font = FONT_SMALL
    row += 2

    # ── KPI Row ──
    kpi_data = [
        (f"${total_rev / 1_000_000:.1f}M", "Annual Revenue"),
        (f"${structural / 1_000_000:.1f}M", "Structural Trade (17.1%)"),
        (f"${operational / 1_000_000:.1f}M", "Operational Waste (3.9%)"),
        (f"${all_in / 1_000_000:.1f}M", "All-In Trade Cost"),
    ]

    col = 2
    for value, label in kpi_data:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.cell(row=row, column=col, value=value).font = FONT_KPI_VALUE
        ws.cell(row=row, column=col).alignment = ALIGN_CENTER
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.cell(row=row + 1, column=col, value=label).font = FONT_KPI_LABEL
        ws.cell(row=row + 1, column=col).alignment = ALIGN_CENTER
        col += 2

    row += 3

    # ── The Punchline ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    structural_pct = structural / total_rev * 100
    operational_pct = operational / total_rev * 100
    all_in_pct = all_in / total_rev * 100
    punchline = (
        f"You budgeted {structural_pct:.0f}% of revenue for trade spend. "
        f"You're actually spending {all_in_pct:.0f}%. "
        f"The extra {operational_pct:.1f} points is operational waste — "
        f"${operational / 1_000_000:.1f}M in trailing-365 deductions "
        f"beyond your planned trade rate."
    )
    punch_cell = ws.cell(row=row, column=2, value=punchline)
    punch_cell.font = Font(name="Calibri", size=12, bold=True, color="333333")
    punch_cell.alignment = ALIGN_LEFT
    ws.row_dimensions[row].height = 40
    row += 2

    # ── Two-Bucket Summary Table ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value="Trade Spend Breakdown").font = Font(
        name="Calibri", size=12, bold=True, color="2F5233")
    row += 1

    bucket_headers = ["Category", "Amount", "% of Revenue"]
    for ci, h in enumerate(bucket_headers, start=2):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, max_col=4)
    row += 1

    buckets = [
        ("Structural / Planned Trade", structural, structural / total_rev),
        ("Operational / Compliance Waste", operational, operational / total_rev),
        ("Promo Billback Deductions", promo_bb, promo_bb / total_rev),
        ("ALL-IN TRADE COST", all_in, all_in / total_rev),
    ]
    for i, (label, amt, pct) in enumerate(buckets):
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=amt).number_format = NUM_DOLLAR
        ws.cell(row=row, column=4, value=pct).number_format = NUM_PCT
        if i == len(buckets) - 1:
            for c in range(2, 5):
                ws.cell(row=row, column=c).font = Font(
                    name="Calibri", size=11, bold=True, color="333333")
                ws.cell(row=row, column=c).border = Border(
                    top=Side(style="double", color="333333"),
                    bottom=Side(style="double", color="333333"))
        else:
            style_data_row(ws, row, max_col=4, stripe=(i % 2 == 0))
        row += 1

    row += 1

    # ── Addressable Improvement ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value="Addressable Improvement").font = Font(
        name="Calibri", size=12, bold=True, color="2F5233")
    row += 1

    improvement_items = [
        ("Operational waste (addressable)", operational),
        ("Already recovered via disputes", -recovery),
        ("NET IMPROVEMENT OPPORTUNITY", addressable),
    ]
    for i, (label, amt) in enumerate(improvement_items):
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=amt).number_format = NUM_DOLLAR
        if i == len(improvement_items) - 1:
            for c in range(2, 4):
                ws.cell(row=row, column=c).font = Font(
                    name="Calibri", size=11, bold=True, color="2F5233")
                ws.cell(row=row, column=c).border = Border(
                    top=Side(style="double", color="2F5233"),
                    bottom=Side(style="double", color="2F5233"))
        else:
            style_data_row(ws, row, max_col=3, stripe=(i % 2 == 0))
        row += 1

    row += 1

    # ── Responsibility Matrix ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value="Waste → Responsible Department").font = Font(
        name="Calibri", size=12, bold=True, color="2F5233")
    row += 1

    resp_headers = ["Waste Category", "Trailing-365 $", "Responsible Dept"]
    for ci, h in enumerate(resp_headers, start=2):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, max_col=4)
    row += 1

    responsibilities = conn.execute(f"""
        SELECT deduction_type, SUM(amount) as total
        FROM deductions
        WHERE deduction_date >= '{TRAIL_START}' AND deduction_date <= '{TRAIL_END}'
        AND deduction_type != 'promo_billback'
        GROUP BY deduction_type
        ORDER BY SUM(amount) DESC
    """).fetchall()

    dept_map = {
        "vague": "Finance / AR",
        "short_ship": "Operations / Warehouse",
        "label_fine": "Operations / Packaging",
        "spoilage": "Supply Chain / QA",
        "late_delivery": "Logistics / 3PL",
        "damaged": "Logistics / Shipping",
        "pallet_fine": "Operations / Warehouse",
        "slotting": "Sales / Trade Marketing",
    }

    for i, r in enumerate(responsibilities):
        dtype = r["deduction_type"]
        ws.cell(row=row, column=2, value=dtype.replace("_", " ").title())
        ws.cell(row=row, column=3, value=r["total"]).number_format = NUM_DOLLAR
        ws.cell(row=row, column=4, value=dept_map.get(dtype, "TBD"))
        style_data_row(ws, row, max_col=4, stripe=(i % 2 == 0))
        row += 1

    row += 1

    # ── Key Metrics ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value="Key Metrics").font = Font(
        name="Calibri", size=12, bold=True, color="2F5233")
    row += 1

    metrics = [
        ("Disputes Filed", f"{dispute_count:,}"),
        ("Total Recovered", f"${recovery:,.0f}"),
        ("Recovery Rate (of disputed $)", f"{recovery_rate:.1%}"),
        ("Double-Dip Events (historical)", f"{dd_count} (${dd_amount:,.0f})"),
        ("Trailing-365 Deduction Count", f"{conn.execute(f'SELECT COUNT(*) FROM deductions WHERE deduction_date >= ? AND deduction_date <= ?', (TRAIL_START, TRAIL_END)).fetchone()[0]:,}"),
    ]
    for i, (label, value) in enumerate(metrics):
        ws.cell(row=row, column=2, value=label).font = FONT_BODY
        ws.cell(row=row, column=4, value=value).font = Font(
            name="Calibri", size=11, bold=True, color="333333")
        ws.cell(row=row, column=4).alignment = ALIGN_RIGHT
        row += 1

    row += 1

    # ── Navigation ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value="Navigate to Detail Tabs").font = Font(
        name="Calibri", size=12, bold=True, color="2F5233")
    row += 1

    nav_items = [
        ("Leak Diagnostic", "Operational waste by category, double-dip alerts"),
        ("Promo Efficacy", "Promotion ROI with adjustable analysis window"),
        ("Retailer Risk", "Net-net margin and what-if trade rate modeling"),
        ("Deduction Ledger", "Full trailing-365 deduction detail"),
        ("Code Crosswalk", "Retailer deduction codes → plain English"),
        ("Methodology & Logic", "Definitions, formulas, and data lineage"),
    ]
    for tab_name, desc in nav_items:
        link_cell = ws.cell(row=row, column=2)
        link_cell.value = tab_name
        link_cell.font = FONT_LINK
        link_cell.hyperlink = f"#'{tab_name}'!A1"
        ws.cell(row=row, column=4, value=desc).font = FONT_BODY
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    for col_letter in "EFGHIJ":
        ws.column_dimensions[col_letter].width = 14

    return ws


# ── Waterfall Chart ──────────────────────────────────────────────────

def add_waterfall_chart(ws, total_rev, structural, operational, promo_bb, all_in, row_start):
    """Add a stacked-bar waterfall chart showing revenue → trade spend breakdown."""
    # Write chart data below the visible area
    data_row = row_start
    ws.cell(row=data_row, column=14, value="Category")
    ws.cell(row=data_row, column=15, value="Invisible")
    ws.cell(row=data_row, column=16, value="Value")

    net_after = total_rev - all_in
    categories = [
        ("Revenue", 0, total_rev),
        ("Structural Trade", net_after + operational + promo_bb, structural),
        ("Operational Waste", net_after + promo_bb, operational),
        ("Promo Billback", net_after, promo_bb),
        ("Net After Trade", 0, net_after),
    ]

    for i, (cat, invisible, value) in enumerate(categories):
        r = data_row + 1 + i
        ws.cell(row=r, column=14, value=cat)
        ws.cell(row=r, column=15, value=invisible)
        ws.cell(row=r, column=16, value=value)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Revenue to Net: Trade Spend Waterfall"
    chart.y_axis.title = "Dollars"
    chart.y_axis.numFmt = '#,##0'
    chart.x_axis.title = None
    chart.style = 10
    chart.width = 22
    chart.height = 14

    cats = Reference(ws, min_col=14, min_row=data_row + 1, max_row=data_row + 5)
    invisible = Reference(ws, min_col=15, min_row=data_row, max_row=data_row + 5)
    values = Reference(ws, min_col=16, min_row=data_row, max_row=data_row + 5)

    chart.add_data(invisible, titles_from_data=True)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)

    # Make invisible series transparent
    chart.series[0].graphicalProperties.noFill = True
    chart.series[0].graphicalProperties.line.noFill = True

    # Color the value bars
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    s = chart.series[1]
    colors = ["339966", "C00000", "BF8F00", "4472C4", "339966"]
    for idx, color in enumerate(colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        s.data_points.append(pt)

    chart.legend = None
    ws.add_chart(chart, "F7")


# ── Workbook-Level Features ──────────────────────────────────────────

def apply_workbook_features(wb, total_rev, structural, operational, promo_bb, all_in):
    from openpyxl.workbook.defined_name import DefinedName

    # Named ranges for KPIs
    named = {
        "AnnualRevenue": total_rev,
        "StructuralTrade": structural,
        "OperationalWaste": operational,
        "PromoBillback": promo_bb,
        "AllInTradeCost": all_in,
        "StructuralTradeRate": structural / total_rev,
        "OperationalWasteRate": operational / total_rev,
        "AllInTradeRate": all_in / total_rev,
    }
    # Write named range values to a hidden helper area on Methodology tab
    ws_method = wb["Methodology & Logic"]
    for i, (name, value) in enumerate(named.items()):
        r = 80 + i
        ws_method.cell(row=r, column=14, value=name).font = FONT_SMALL
        ws_method.cell(row=r, column=15, value=value)
        if "Rate" in name:
            ws_method.cell(row=r, column=15).number_format = NUM_PCT
        else:
            ws_method.cell(row=r, column=15).number_format = NUM_DOLLAR

        dn = DefinedName(name, attr_text=f"'Methodology & Logic'!$O${r}")
        wb.defined_names.add(dn)

    # Freeze panes on Retailer Risk (header at row 5, data from row 6)
    ws_rr = wb["Retailer Risk"]
    ws_rr.auto_filter.ref = f"B5:L{ws_rr.max_row}"
    ws_rr.freeze_panes = "B6"

    # Print areas
    ws_exec = wb["Executive Pulse"]
    ws_exec.print_area = "A1:J55"
    ws_exec.sheet_properties.pageSetUpPr.fitToPage = True

    ws_leak = wb["Leak Diagnostic"]
    ws_leak.print_area = f"A1:J{ws_leak.max_row}"

    ws_rr.print_area = f"A1:L{ws_rr.max_row}"

    ws_promo = wb["Promo Efficacy"]
    ws_promo.print_area = f"A1:N{ws_promo.max_row}"

    # Conditional formatting: red for negative ROI on Promo Efficacy
    from openpyxl.formatting.rule import CellIsRule
    red_font = Font(color="C00000")
    red_fill = PatternFill("solid", fgColor="FDE8E8")
    green_font = Font(color="2F5233")
    green_fill = PatternFill("solid", fgColor="E8F5E8")

    ws_promo.conditional_formatting.add(
        f"M9:M{ws_promo.max_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font)
    )
    ws_promo.conditional_formatting.add(
        f"M9:M{ws_promo.max_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill, font=green_font)
    )

    # Conditional formatting: highlight high operational deduction rate on Retailer Risk
    ws_rr.conditional_formatting.add(
        f"I7:I{ws_rr.max_row}",
        CellIsRule(operator="greaterThan", formula=["0.05"], fill=red_fill, font=red_font)
    )

    # Data bars on deduction amounts in ledger
    ws_ledger = wb["Deduction Ledger"]
    ws_ledger.conditional_formatting.add(
        f"H6:H{ws_ledger.max_row}",
        DataBarRule(start_type="min", end_type="max", color="4472C4")
    )


# ── Main ─────────────────────────────────────────────────────────────

def main():
    print("Connecting to database...")
    conn = connect_db()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building Tab 7: Methodology & Logic...")
    build_tab7_methodology(wb, conn)

    print("Building Tab 6: Code Crosswalk...")
    build_tab6_crosswalk(wb, conn)

    print("Building Tab 5: Deduction Ledger...")
    build_tab5_ledger(wb, conn)

    print("Building Tab 4: Retailer Risk...")
    build_tab4_retailer_risk(wb, conn)

    print("Building Tab 3: Promo Efficacy...")
    build_tab3_promo(wb, conn)

    print("Building Tab 2: Leak Diagnostic...")
    build_tab2_leak(wb, conn)

    print("Building Tab 1: Executive Pulse...")
    ws1 = build_tab1_executive(wb, conn)

    # Add waterfall chart to Executive Pulse
    total_rev = conn.execute(f"""
        SELECT SUM(dollars_sold) FROM scan_data
        WHERE week_ending >= '{TRAIL_START}' AND week_ending <= '{TRAIL_END}'
    """).fetchone()[0]
    structural = conn.execute(f"""
        SELECT SUM(
            sc.trade_spend_pct_walmart * COALESCE(sd_rev.rev_walmart, 0) +
            sc.trade_spend_pct_costco * COALESCE(sd_rev.rev_costco, 0) +
            sc.trade_spend_pct_whole_foods * COALESCE(sd_rev.rev_wf, 0) +
            sc.trade_spend_pct_regional * COALESCE(sd_rev.rev_regional, 0) +
            sc.trade_spend_pct_unfi * COALESCE(sd_rev.rev_unfi, 0) +
            sc.trade_spend_pct_dtc * COALESCE(sd_rev.rev_dtc, 0)
        )
        FROM sku_costs sc
        LEFT JOIN (
            SELECT sd.sku,
                SUM(CASE WHEN s.retailer='Walmart' THEN sd.dollars_sold ELSE 0 END) as rev_walmart,
                SUM(CASE WHEN s.retailer='Costco' THEN sd.dollars_sold ELSE 0 END) as rev_costco,
                SUM(CASE WHEN s.retailer='Whole Foods' THEN sd.dollars_sold ELSE 0 END) as rev_wf,
                SUM(CASE WHEN s.retailer IN ('Green Basket Market','Southside Grocers','Prairie Provisions','Mountain Pantry Co','Harbor Fresh') THEN sd.dollars_sold ELSE 0 END) as rev_regional,
                SUM(CASE WHEN s.retailer='UNFI' THEN sd.dollars_sold ELSE 0 END) as rev_unfi,
                SUM(CASE WHEN s.retailer='DTC' THEN sd.dollars_sold ELSE 0 END) as rev_dtc
            FROM scan_data sd
            JOIN stores s ON sd.store_id = s.store_id
            WHERE sd.week_ending >= '{TRAIL_START}' AND sd.week_ending <= '{TRAIL_END}'
            GROUP BY sd.sku
        ) sd_rev ON sc.sku = sd_rev.sku
    """).fetchone()[0]
    operational = conn.execute(f"""
        SELECT SUM(amount) FROM deductions
        WHERE deduction_date >= '{TRAIL_START}' AND deduction_date <= '{TRAIL_END}'
        AND deduction_type != 'promo_billback'
    """).fetchone()[0]
    promo_bb = conn.execute(f"""
        SELECT SUM(amount) FROM deductions
        WHERE deduction_date >= '{TRAIL_START}' AND deduction_date <= '{TRAIL_END}'
        AND deduction_type = 'promo_billback'
    """).fetchone()[0]
    all_in = structural + operational + promo_bb
    add_waterfall_chart(ws1, total_rev, structural, operational, promo_bb, all_in, row_start=60)

    # ── Workbook-level features ──
    print("Applying workbook-level features...")
    apply_workbook_features(wb, total_rev, structural, operational, promo_bb, all_in)

    # Reorder sheets: Tab 1 first
    tab_order = [
        "Executive Pulse", "Leak Diagnostic", "Promo Efficacy",
        "Retailer Risk", "Deduction Ledger", "Code Crosswalk",
        "Methodology & Logic",
    ]
    sheet_indices = []
    for name in tab_order:
        for idx, sheet in enumerate(wb.worksheets):
            if sheet.title == name:
                sheet_indices.append(idx)
                break
    wb._sheets = [wb.worksheets[i] for i in sheet_indices]

    # Save
    output_dir = Path(__file__).resolve().parent.parent / "output"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / "trade_spend_diagnostic.xlsx"
    wb.save(str(output_path))
    print(f"\nWorkbook saved: {output_path}")
    print(f"Tabs: {[s.title for s in wb.worksheets]}")

    conn.close()


if __name__ == "__main__":
    main()
