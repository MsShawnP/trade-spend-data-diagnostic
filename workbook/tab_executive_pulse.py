"""Tab 1: Executive Pulse — the CEO punchline."""

from datetime import date

from workbook.db import connect

from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from workbook.channel_mapping import CHANNEL_RATE_COLS, RETAILER_TO_CHANNEL
from workbook.deduction_taxonomy import get_taxonomy
from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    FONT_BODY,
    FONT_HEADER,
    FONT_KPI_LABEL,
    FONT_KPI_VALUE,
    FONT_NAV,
    FONT_SECTION,
    FONT_SMALL,
    NUM_FMT_DOLLAR,
    NUM_FMT_PCT,
    TABLE_STYLE,
)

CATEGORY_TO_DEPT = {
    "slotting": ("Sales", "Negotiated trade terms"),
    "short_ship": ("Operations", "Fulfillment accuracy"),
    "late_delivery": ("Operations", "Logistics/carrier management"),
    "damaged": ("Operations", "Packaging/handling"),
    "pallet_fine": ("Operations", "Compliance with retailer specs"),
    "label_fine": ("Operations", "Labeling/packaging compliance"),
    "spoilage": ("Operations", "Shelf-life/inventory management"),
    "vague": ("Finance", "Unclear codes — investigate"),
}

BENCHMARK_BAND = (0.19, 0.23)
BENCHMARK_NOTE = (
    "Natural/specialty CPG category average"
    " (Cadent Consulting, Booz & Company trade promotion benchmarks)"
)

NAV_TABS = [
    "Executive Pulse",
    "Leak Diagnostic",
    "Promo Efficacy",
    "Retailer Risk",
    "Deduction Ledger",
    "Deduction Code Crosswalk",
    "Methodology & Logic",
]


def _query_metrics(database_url: str) -> dict:
    conn = connect()

    weeks = conn.execute(
        "SELECT DISTINCT week_ending FROM stg_scan_data ORDER BY week_ending DESC LIMIT 52"
    ).fetchall()
    oldest_week = weeks[-1][0]

    revenue = conn.execute(
        "SELECT SUM(dollars_sold) FROM stg_scan_data WHERE week_ending >= %s",
        (oldest_week,),
    ).fetchone()[0]

    channel_rev = conn.execute("""
        SELECT s.retailer, SUM(sd.dollars_sold)
        FROM stg_scan_data sd
        JOIN stg_stores s ON sd.store_id = s.store_id
        WHERE sd.week_ending >= %s
        GROUP BY s.retailer
    """, (oldest_week,)).fetchall()

    rates = {}
    for channel, col in CHANNEL_RATE_COLS.items():
        rates[channel] = conn.execute(f"SELECT AVG({col}) FROM stg_sku_costs").fetchone()[0]

    channel_trade = 0.0
    for retailer, rev in channel_rev:
        channel_trade += rev * rates.get(retailer, rates["Regional"])

    max_scan = weeks[0][0]
    deductions_by_type = conn.execute("""
        SELECT deduction_type, COUNT(*), SUM(amount)
        FROM stg_deductions
        WHERE deduction_date > (%s::date - interval '365 days')::date AND deduction_date <= %s
          AND deduction_type != 'promo_billback'
        GROUP BY deduction_type
        ORDER BY SUM(amount) DESC
    """, (max_scan, max_scan)).fetchall()

    operational_waste = sum(r[2] for r in deductions_by_type)
    addressable_waste = sum(
        r[2] for r in deductions_by_type if get_taxonomy(r[0])["addressable"]
    )

    # Waste by retailer → aggregate to channel
    waste_by_retailer = conn.execute("""
        SELECT retailer_id, SUM(amount)
        FROM stg_deductions
        WHERE deduction_date > (%s::date - interval '365 days')::date AND deduction_date <= %s
          AND deduction_type != 'promo_billback'
        GROUP BY retailer_id
    """, (max_scan, max_scan)).fetchall()

    # Map retailer slugs to display names for channel lookup
    retailer_names = dict(conn.execute(
        "SELECT LOWER(REPLACE(name, ' ', '_')), name FROM stg_retailers"
    ).fetchall())

    waste_by_channel: dict[str, float] = {}
    for slug, amount in waste_by_retailer:
        display_name = retailer_names.get(slug, slug)
        channel = RETAILER_TO_CHANNEL.get(display_name, "Other")
        waste_by_channel[channel] = waste_by_channel.get(channel, 0) + amount

    dispute_stats = conn.execute("""
        SELECT COUNT(*), SUM(recovered_amount)
        FROM stg_disputes
    """).fetchone()

    total_disputed = conn.execute("""
        SELECT SUM(d.amount) FROM stg_deductions d
        JOIN stg_disputes dis ON dis.deduction_id = d.deduction_id
    """).fetchone()[0] or 0

    # Monthly waste — exclude partial months (<20 days of data) to avoid trend bias
    monthly_waste_raw = conn.execute("""
        SELECT
            to_char(deduction_date, 'YYYY-MM') as month,
            SUM(amount) as monthly_total,
            COUNT(DISTINCT deduction_date) as active_days
        FROM stg_deductions
        WHERE deduction_date > (%s::date - interval '365 days')::date AND deduction_date <= %s
          AND deduction_type != 'promo_billback'
        GROUP BY to_char(deduction_date, 'YYYY-MM')
        ORDER BY month
    """, (max_scan, max_scan)).fetchall()
    monthly_waste = [(m, t) for m, t, days in monthly_waste_raw if days >= 20]

    conn.close()

    return {
        "revenue": revenue,
        "structural_trade": channel_trade,
        "operational_waste": operational_waste,
        "addressable_waste": addressable_waste,
        "deductions_by_type": deductions_by_type,
        "dispute_count": dispute_stats[0],
        "total_recovered": dispute_stats[1],
        "total_disputed": total_disputed,
        "waste_by_channel": waste_by_channel,
        "oldest_week": oldest_week,
        "max_scan": max_scan,
        "monthly_waste": monthly_waste,
    }


def build_executive_pulse(ws: Worksheet, database_url: str) -> None:
    metrics = _query_metrics(database_url)

    revenue = metrics["revenue"]
    structural = metrics["structural_trade"]
    waste = metrics["operational_waste"]
    all_in = structural + waste
    net_after = revenue - all_in

    structural_rate = structural / revenue
    waste_rate = waste / revenue
    all_in_rate = all_in / revenue

    recovery_rate = metrics["total_recovered"] / metrics["total_disputed"] if metrics["total_disputed"] else 0
    total_recovered = metrics["total_recovered"]

    ws.sheet_view.showGridLines = False

    col_widths = [3, 32, 16, 14, 32, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Header block (rows 1-3) ---
    ws.merge_cells("B1:F1")
    ws["B1"] = "Cinderhaven Provisions"
    ws["B1"].font = FONT_HEADER

    ws.merge_cells("B2:F2")
    ws["B2"] = "Executive Pulse — Trade Spend Diagnostic"
    ws["B2"].font = FONT_SECTION

    ws.merge_cells("B3:F3")
    ws["B3"] = f"Trailing 52 weeks ({metrics['oldest_week']} to {metrics['max_scan']})  |  Built {date.today().isoformat()}"
    ws["B3"].font = FONT_SMALL

    addressable = metrics["addressable_waste"]

    # --- KPI row (rows 5-7) ---
    kpis = [
        ("All-In Trade Rate", all_in_rate, NUM_FMT_PCT),
        ("Planned Trade Rate", structural_rate, NUM_FMT_PCT),
        ("Operational Waste", waste_rate, NUM_FMT_PCT),
        ("Addressable Waste", addressable, NUM_FMT_DOLLAR),
    ]
    for col_idx, (label, value, fmt) in enumerate(kpis, 2):
        cell_val = ws.cell(row=5, column=col_idx, value=value)
        cell_val.font = FONT_KPI_VALUE
        cell_val.number_format = fmt
        cell_val.alignment = ALIGN_CENTER

        cell_lbl = ws.cell(row=6, column=col_idx, value=label)
        cell_lbl.font = FONT_KPI_LABEL
        cell_lbl.alignment = ALIGN_CENTER

    # --- Benchmark annotation (column F, next to KPIs) ---
    bench_val = ws.cell(
        row=5, column=6,
        value=f"{BENCHMARK_BAND[0]*100:.0f}–{BENCHMARK_BAND[1]*100:.0f}%",
    )
    bench_val.font = Font(name="Calibri", size=12, bold=True, color="808080")
    bench_val.alignment = ALIGN_CENTER
    bench_lbl = ws.cell(row=6, column=6, value="Industry Range")
    bench_lbl.font = Font(name="Calibri", size=9, color="808080")
    bench_lbl.alignment = ALIGN_CENTER

    top_channels = sorted(metrics["waste_by_channel"].items(), key=lambda x: x[1], reverse=True)[:2]
    ch_note = ""
    if top_channels:
        parts = [f"{ch} (${amt:,.0f})" for ch, amt in top_channels]
        ch_note = f" Top waste channels: {', '.join(parts)}."

    ws.merge_cells("B7:F7")
    ws["B7"] = (
        f"You budgeted 17%. You're at {all_in_rate*100:.0f}%"
        f" (industry: {BENCHMARK_BAND[0]*100:.0f}–{BENCHMARK_BAND[1]*100:.0f}%)."
        f" The gap is {waste_rate*100:.0f} points of operational waste.{ch_note}"
    )
    ws["B7"].font = Font(name="Calibri", size=11, italic=True)
    ws["B7"].alignment = ALIGN_LEFT

    section_sep = Border(bottom=Side(style="thin", color="CCCCCC"))
    for col in range(2, 7):
        ws.cell(row=8, column=col).border = section_sep

    # --- Hidden source data (rows 10-14) for named ranges ---
    ws.cell(row=9, column=2, value="Waterfall: Revenue to Net-After-Trade").font = FONT_SECTION

    chart_start = 10
    headers = ["Category", "Base", "Value"]
    for c, h in enumerate(headers, 2):
        ws.cell(row=chart_start, column=c, value=h).font = FONT_BODY

    waterfall_data = [
        ("Revenue", 0, revenue),
        ("Structural Trade", net_after + waste, structural),
        ("Operational Waste", net_after, waste),
        ("Net After Trade", 0, net_after),
    ]
    for i, (cat, base, val) in enumerate(waterfall_data):
        r = chart_start + 1 + i
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=base).number_format = NUM_FMT_DOLLAR
        ws.cell(row=r, column=4, value=val).number_format = NUM_FMT_DOLLAR

    for r in range(chart_start, chart_start + 5):
        ws.row_dimensions[r].hidden = True

    # --- In-cell waterfall (rows 16-19) ---
    waterfall_visible = [
        ("Revenue", f"=D{chart_start + 1}", "2F5496"),
        ("Structural Trade", f"=D{chart_start + 2}", "ED7D31"),
        ("Operational Waste", f"=D{chart_start + 3}", "C00000"),
        ("Net After Trade", f"=D{chart_start + 4}", "2F5496"),
    ]
    for i, (label, formula, color) in enumerate(waterfall_visible):
        r = 16 + i
        ws.cell(row=r, column=2, value=label).font = FONT_BODY
        val_cell = ws.cell(row=r, column=4, value=formula)
        val_cell.number_format = NUM_FMT_DOLLAR
        val_cell.alignment = ALIGN_RIGHT

        ws.conditional_formatting.add(
            f"D{r}",
            DataBarRule(
                start_type="num", start_value=0,
                end_type="num", end_value=revenue,
                color=color,
            ),
        )

    for col in range(2, 7):
        ws.cell(row=21, column=col).border = section_sep

    # --- Addressable Improvement (rows 22-30) ---
    ws.cell(row=22, column=2, value="Addressable Improvement").font = FONT_SECTION

    imp_header_row = 23
    imp_headers = ["Metric", "Value"]
    for c, h in enumerate(imp_headers, 2):
        cell = ws.cell(row=imp_header_row, column=c, value=h)
        cell.font = FONT_BODY
        cell.alignment = ALIGN_CENTER

    improvement_rows = [
        ("Operational waste (trailing 365d)", waste, NUM_FMT_DOLLAR),
        ("Current recovery rate", recovery_rate, NUM_FMT_PCT),
        ("Currently recovered", total_recovered, NUM_FMT_DOLLAR),
        ("At 30% recovery", waste * 0.30, NUM_FMT_DOLLAR),
        ("At 50% recovery", waste * 0.50, NUM_FMT_DOLLAR),
        ("Incremental at 30% vs. current", waste * 0.30 - total_recovered, NUM_FMT_DOLLAR),
        ("Incremental at 50% vs. current", waste * 0.50 - total_recovered, NUM_FMT_DOLLAR),
    ]
    for i, (label, val, fmt) in enumerate(improvement_rows):
        r = imp_header_row + 1 + i
        ws.cell(row=r, column=2, value=label).font = FONT_BODY
        c_val = ws.cell(row=r, column=3, value=val)
        c_val.font = FONT_BODY
        c_val.number_format = fmt
        c_val.alignment = ALIGN_RIGHT

    imp_end_row = imp_header_row + len(improvement_rows)

    imp_table = Table(displayName="tbl_AddressableImprovement", ref=f"B{imp_header_row}:C{imp_end_row}")
    imp_table.tableStyleInfo = TABLE_STYLE
    ws.add_table(imp_table)

    for col in range(2, 7):
        ws.cell(row=imp_end_row + 1, column=col).border = section_sep

    # --- Responsibility Matrix (rows 32+) ---
    resp_title_row = imp_end_row + 2
    ws.cell(row=resp_title_row, column=2, value="Responsibility Matrix — Who Owns the Waste").font = FONT_SECTION

    resp_header_row = resp_title_row + 1
    resp_headers = ["Deduction Type", "Amount", "Owner", "Root Cause"]
    for c, h in enumerate(resp_headers, 2):
        cell = ws.cell(row=resp_header_row, column=c, value=h)
        cell.font = FONT_BODY
        cell.alignment = ALIGN_CENTER

    for i, (dtype, count, amount) in enumerate(metrics["deductions_by_type"]):
        r = resp_header_row + 1 + i
        dept, cause = CATEGORY_TO_DEPT.get(dtype, ("Unclassified", ""))
        ws.cell(row=r, column=2, value=dtype.replace("_", " ").title())
        c_amt = ws.cell(row=r, column=3, value=amount)
        c_amt.number_format = NUM_FMT_DOLLAR
        c_amt.alignment = ALIGN_RIGHT
        ws.cell(row=r, column=4, value=dept)
        ws.cell(row=r, column=5, value=cause)

    resp_end_row = resp_header_row + len(metrics["deductions_by_type"])

    resp_table = Table(displayName="tbl_ResponsibilityMatrix", ref=f"B{resp_header_row}:E{resp_end_row}")
    resp_table.tableStyleInfo = TABLE_STYLE
    ws.add_table(resp_table)

    # --- Navigation hyperlinks ---
    nav_row = resp_end_row + 2
    ws.cell(row=nav_row, column=2, value="Navigate to:").font = FONT_SECTION

    nav_row += 1
    for i, tab_name in enumerate(NAV_TABS):
        cell = ws.cell(row=nav_row, column=2 + i, value=tab_name)
        cell.font = FONT_NAV
        cell.hyperlink = f"#'{tab_name}'!A1"

    # --- Instructional callout ---
    callout_row = nav_row + 2
    ws.merge_cells(f"B{callout_row}:F{callout_row}")
    callout = ws.cell(
        row=callout_row, column=2,
        value=(
            "This workbook summarizes trailing-12-month trade spend for Cinderhaven Provisions. "
            "Green tabs are analysis. Blue tab is the full deduction ledger. "
            "Gray tabs are reference. Yellow cells are adjustable inputs."
        ),
    )
    callout.font = FONT_SMALL
    callout.alignment = ALIGN_LEFT

    # --- Waste trend analysis (text-based — openpyxl charts don't render) ---
    monthly = metrics["monthly_waste"]
    if len(monthly) >= 6:
        mid = len(monthly) // 2
        h1_avg = sum(t for _, t in monthly[:mid]) / mid
        h2_avg = sum(t for _, t in monthly[mid:]) / (len(monthly) - mid)
        mo_avg = sum(t for _, t in monthly) / len(monthly)
        pct_change = (h2_avg - h1_avg) / h1_avg * 100

        if abs(pct_change) < 5:
            direction = "stable"
        elif pct_change > 0:
            direction = "rising"
        else:
            direction = "declining"

        trend_text = (
            f"12-month waste trend: {direction} at ${mo_avg:,.0f}/mo avg."
            f" H2 {'up' if pct_change > 0 else 'down'}"
            f" {abs(pct_change):.0f}% vs H1."
        )
    else:
        mo_avg = sum(t for _, t in monthly) / len(monthly) if monthly else 0
        trend_text = f"Waste run-rate: ${mo_avg:,.0f}/mo avg."

    ws.merge_cells("F16:G19")
    trend_cell = ws.cell(row=16, column=6, value=trend_text)
    trend_cell.font = Font(name="Calibri", size=10, italic=True, color="C00000")
    trend_cell.alignment = Alignment(vertical="top", wrap_text=True)
