"""Tab 7: Methodology & Logic — self-contained documentation of all calculations."""

import contextlib
import os
import sqlite3
from datetime import date
from pathlib import Path

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from workbook.queries import fetch_channel_rates, get_trailing_bounds
from workbook.styles import FONT_BODY, FONT_HEADER, FONT_SECTION, FONT_SMALL, LONDON_20, SANS

_LABEL_FONT = Font(name=SANS, size=11, bold=True, color=LONDON_20)
_WRAP = Alignment(vertical="top", wrap_text=True)


def _write_section(ws: Worksheet, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title).font = FONT_SECTION
    return row + 1


def _write_body(ws: Worksheet, row: int, text: str) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_BODY
    cell.alignment = _WRAP
    return row + 1


def _write_pair(ws: Worksheet, row: int, label: str, text: str) -> int:
    c1 = ws.cell(row=row, column=1, value=label)
    c1.font = _LABEL_FONT
    c1.alignment = Alignment(vertical="top")
    c2 = ws.cell(row=row, column=2, value=text)
    c2.font = FONT_BODY
    c2.alignment = _WRAP
    return row + 1


def _query_methodology_numbers(db_path: Path) -> dict:
    with contextlib.closing(sqlite3.connect(db_path)) as conn:
        oldest_week, max_scan = get_trailing_bounds(conn)

        revenue = conn.execute(
            "SELECT SUM(dollars_sold) FROM scan_data WHERE week_ending >= ?",
            (oldest_week,),
        ).fetchone()[0]
        if not revenue:
            raise ValueError("No revenue data found in scan_data for the trailing window")

        channel_rev = conn.execute(
            "SELECT s.retailer, SUM(sd.dollars_sold) FROM scan_data sd "
            "JOIN stores s ON sd.store_id = s.store_id "
            "WHERE sd.week_ending >= ? GROUP BY s.retailer",
            (oldest_week,),
        ).fetchall()
        rates, regional_rate = fetch_channel_rates(conn)
        structural = sum(r * rates.get(ch, regional_rate) for ch, r in channel_rev)

        ded_sums = conn.execute(
            "SELECT "
            "SUM(CASE WHEN deduction_type != 'promo_billback' THEN amount ELSE 0 END), "
            "SUM(CASE WHEN deduction_type = 'promo_billback' THEN amount ELSE 0 END) "
            "FROM deductions "
            "WHERE deduction_date > date(?, '-365 days') AND deduction_date <= ?",
            (max_scan, max_scan),
        ).fetchone()
        waste = ded_sums[0] or 0
        pb = ded_sums[1] or 0

        total_ded = conn.execute("SELECT COUNT(*) FROM deductions").fetchone()[0]
        ded_range = conn.execute(
            "SELECT MIN(deduction_date), MAX(deduction_date) FROM deductions"
        ).fetchone()

        promo_row = conn.execute(
            "SELECT COUNT(*), COUNT(DISTINCT promo_id), COALESCE(SUM(promo_cost), 0) FROM promotions"
        ).fetchone()
        promo_rows, promo_distinct, promo_cost_sum = promo_row

        code_row = conn.execute(
            "SELECT COUNT(*), SUM(CASE WHEN is_published = 1 THEN 1 ELSE 0 END) FROM deduction_codes"
        ).fetchone()
        code_total = code_row[0]
        code_published = code_row[1]
        code_inferred = code_total - code_published

        dispute_count = conn.execute("SELECT COUNT(*) FROM disputes").fetchone()[0]
        recovered = conn.execute("SELECT SUM(recovered_amount) FROM disputes").fetchone()[0] or 0
        disputed_total = conn.execute(
            "SELECT SUM(d.amount) FROM deductions d "
            "JOIN disputes dis ON dis.deduction_id = d.deduction_id"
        ).fetchone()[0] or 0

        scan_range = conn.execute(
            "SELECT MIN(week_ending), MAX(week_ending), COUNT(DISTINCT week_ending) FROM scan_data"
        ).fetchone()

        table_count = conn.execute(
            "SELECT COUNT(*) FROM sqlite_master WHERE type='table'"
        ).fetchone()[0]

        db_size_mb = os.path.getsize(db_path) / (1024 * 1024)

    all_in = structural + waste
    return {
        "revenue": revenue,
        "structural": structural,
        "structural_pct": structural / revenue * 100,
        "waste": waste,
        "waste_pct": waste / revenue * 100,
        "all_in": all_in,
        "all_in_pct": all_in / revenue * 100,
        "pb": pb,
        "total_ded": total_ded,
        "ded_start": ded_range[0],
        "ded_end": ded_range[1],
        "promo_rows": promo_rows,
        "promo_distinct": promo_distinct,
        "promo_cost_sum": promo_cost_sum,
        "code_total": code_total,
        "code_published": code_published,
        "code_inferred": code_inferred,
        "dispute_count": dispute_count,
        "recovered": recovered,
        "disputed_total": disputed_total,
        "recovery_pct": recovered / disputed_total * 100 if disputed_total else 0,
        "scan_start": scan_range[0],
        "scan_end": scan_range[1],
        "scan_weeks": scan_range[2],
        "table_count": table_count,
        "db_size_mb": db_size_mb,
    }


def build_methodology(ws: Worksheet, db_path: Path) -> None:
    n = _query_methodology_numbers(db_path)

    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 90

    # --- Header ---
    row = 1
    ws.merge_cells("A1:B1")
    ws["A1"] = "Methodology & Logic"
    ws["A1"].font = FONT_HEADER

    row = 2
    ws["A2"] = f"Build date: {date.today().isoformat()}"
    ws["A2"].font = FONT_SMALL

    # === TWO-BUCKET DEFINITION ===
    row = 4
    row = _write_section(ws, row, "1. Two-Bucket Executive Framing")
    row = _write_body(ws, row,
        "This workbook uses a two-bucket model to present trade spend:")
    row += 1
    row = _write_pair(ws, row, "Bucket 1: Structural Trade",
        "The negotiated rate-card discount embedded in wholesale pricing by channel. "
        "Derived from sku_costs.trade_spend_pct columns multiplied by channel revenue. "
        "This is the planned cost of doing business with each retailer — it exists "
        f"whether or not a single deduction is ever taken. ${n['structural']:,.0f} ({n['structural_pct']:.1f}% of revenue).")
    row = _write_pair(ws, row, "Bucket 2: Operational Waste",
        "Trailing-365-day deductions excluding promo_billback. These are unplanned cash "
        "outflows from compliance failures (label fines, pallet fines), logistics issues "
        "(short ships, late deliveries, damages), spoilage, pricing and labeling errors. "
        f"${n['waste']:,.0f} ({n['waste_pct']:.1f}% of revenue).")
    row += 1
    row = _write_pair(ws, row, "Why not three buckets?",
        "The original brief proposed a third \"promotional\" bucket using off-invoice discounts. "
        "Investigation revealed that off-invoice is a funding mechanism, not a cost category — "
        "including it as a separate bucket double-counts costs already captured in the structural "
        f"trade rate. The promotions table's promo_cost sum (${n['promo_cost_sum']:,.0f}) is too small to constitute "
        f"a meaningful standalone bucket. Two buckets tell a cleaner story: structural trade at {n['structural_pct']:.0f}% "
        f"is competitive; the {n['waste_pct']:.1f}% operational waste layer is where the addressable money is.")
    row = _write_pair(ws, row, "Promo billback exclusion",
        f"Promo_billback deductions (${n['pb']:,.0f} trailing-365) are excluded from the operational "
        "waste bucket because they represent planned promotional activity, not operational failures. "
        "They appear on the Deduction Ledger tab but do not inflate the waste figure.")

    # === DATA LINEAGE ===
    row += 1
    row = _write_section(ws, row, "2. Data Lineage")
    row = _write_body(ws, row,
        "All data originates from the cinderhaven-data SQLite database (cinderhaven_product_master.db), "
        f"built via the cinderhaven-data/build_db.py pipeline. {n['table_count']} tables, {n['db_size_mb']:.1f} MB.")
    row += 1
    row = _write_pair(ws, row, "scan_data",
        "Point-of-sale weekly volumes and dollar sales by SKU and store. "
        f"{n['scan_weeks']} weeks ({n['scan_start']} to {n['scan_end']}). Used for revenue calculations (trailing 52 weeks = "
        "52 most recent distinct week_ending values) and promotion lift analysis.")
    row = _write_pair(ws, row, "sku_costs",
        "Per-SKU cost structure: COGS, wholesale prices by channel, trade spend percentages by channel. "
        "Static reference table. Used for structural trade rate calculation and gross margin derivation.")
    row = _write_pair(ws, row, "deductions",
        f"{n['total_ded']:,} deduction records ({n['ded_start']} to {n['ded_end']}). Each record: retailer, type, amount, date, "
        "codes and status flags. Trailing-365 filter applied for operational waste "
        "calculations. Joined to deduction_codes for translations and to disputes for recovery data.")
    row = _write_pair(ws, row, "promotions",
        f"{n['promo_rows']} promotion rows across {n['promo_distinct']} distinct events. Fields: SKU, retailer, date window, promo type, "
        "promo_cost, funding_mechanism. Coverage: not all promotions have matched POS data. "
        "Limitation: promo calendar may be incomplete — ghost promo analysis identifies deductions "
        "that reference promotions not in this table.")
    row = _write_pair(ws, row, "deduction_codes",
        f"{n['code_total']} retailer-specific code entries mapping raw remittance codes to plain-English descriptions "
        f"and standardized categories. {n['code_published']} verified from vendor guides, {n['code_inferred']} inferred via pattern matching.")
    row = _write_pair(ws, row, "disputes",
        f"{n['dispute_count']:,} dispute records with outcome, recovered amount, filed/closed dates. "
        "Joined to deductions on deduction_id. Recovery rate = total recovered / total disputed dollars.")
    row = _write_pair(ws, row, "stores",
        "Store-to-retailer mapping. Used to aggregate scan_data from store level to retailer/channel level.")

    # === ROI METHODOLOGY ===
    row += 1
    row = _write_section(ws, row, "3. Promotion ROI Methodology")
    row = _write_body(ws, row,
        "Simple pre/during/post volume comparison. Not a causal model.")
    row += 1
    row = _write_pair(ws, row, "Step 1: Baseline",
        "Average weekly unit volume for the SKU at that retailer over the N weeks immediately "
        "preceding the promotion start date, where N = the adjustable window parameter (default 4, range 1–8).")
    row = _write_pair(ws, row, "Step 2: During-period",
        "Average weekly unit volume during the promotion weeks (start_week through end_week).")
    row = _write_pair(ws, row, "Step 3: Incremental volume",
        "(During-period avg − Baseline avg) × promotion duration in weeks.")
    row = _write_pair(ws, row, "Step 4: Incremental revenue",
        "Incremental volume × average selling price (ASP) for that SKU at that retailer, "
        "calculated from scan_data as dollars_sold / units_sold.")
    row = _write_pair(ws, row, "Step 5: ROI",
        "Incremental revenue ÷ promotion cost. Uses actual cost (matched promo_billback deduction) "
        "if available; otherwise uses planned cost from the promotions table and flags it.")
    row += 1
    row = _write_pair(ws, row, "Window parameter",
        "The pre/post window (Tab 3 cell D5) controls how many weeks are used for baseline and "
        "post-period comparison. Larger windows smooth noise but may include other promotions. "
        "Smaller windows are more sensitive but noisier. All ROI formulas reference this cell — "
        "changing it recalculates every promotion's ROI.")
    row = _write_pair(ws, row, "Limitations",
        "This methodology does not adjust for seasonality, does not establish causality, "
        "does not control for distribution changes or out-of-stocks, and does not model "
        "post-promotion dip (pantry-loading). Sophisticated baseline modeling with "
        "causal inference is engagement-level work beyond this diagnostic's scope.")

    # === NET-NET MARGIN ===
    row += 1
    row = _write_section(ws, row, "4. Net-Net Effective Margin Methodology")
    row = _write_body(ws, row,
        "Margin waterfall from gross to effective, calculated per retailer:")
    row += 1
    row = _write_pair(ws, row, "Gross margin",
        "( Wholesale price − COGS ) / Wholesale price. Uses channel-specific wholesale prices "
        "from sku_costs, averaged across all SKUs. Example: Walmart GM = 39.0%, DTC GM = 62.5%.")
    row = _write_pair(ws, row, "After structural trade",
        "Gross margin − structural trade rate. The structural rate is the average "
        "trade_spend_pct for that channel from sku_costs.")
    row = _write_pair(ws, row, "Net-net effective margin",
        "After-structural margin − (operational deductions + promo billback) / revenue. "
        "This is the true margin after all trade costs are accounted for.")
    row = _write_pair(ws, row, "Exclusions",
        "Freight, warehousing, and non-trade SG&A are not included. This is trade margin only, "
        "not net income margin.")

    # === RECOVERY METHODOLOGY ===
    row += 1
    row = _write_section(ws, row, "5. Recovery Rate & Addressable Improvement")
    row = _write_pair(ws, row, "Current recovery rate",
        f"Total recovered dollars (${n['recovered']:,.0f}) ÷ total disputed dollars (${n['disputed_total']:,.0f}) = {n['recovery_pct']:.1f}%. "
        "This counts won (100% recovery) and partial (~49% average recovery) outcomes.")
    row = _write_pair(ws, row, "Target recovery input",
        "Tab 2 cell C41 allows entering a target recovery rate (0–100%). "
        "Recovery at target = operational waste × target rate. "
        "Incremental opportunity = recovery at target − current recovered.")
    row = _write_pair(ws, row, "Interpretation",
        "The addressable improvement assumes all operational waste is disputable. "
        "In practice, some categories (slotting, label fines) have low recoverability. "
        "The recoverability score on Tab 2 provides qualitative guidance on which "
        "categories realistically support higher recovery targets.")

    # === GLOSSARY ===
    row += 1
    row = _write_section(ws, row, "6. Glossary")
    terms = [
        ("Trade spend", "All costs a manufacturer pays to retailers beyond product COGS — includes rate-card discounts, promotional allowances, and compliance deductions."),
        ("Structural trade", "The contractual discount rate embedded in wholesale pricing, applied to every unit sold through that channel."),
        ("Operational waste", "Deductions taken by retailers for operational or compliance issues — not planned promotional activity."),
        ("Off-invoice", "A funding mechanism where the promotional discount is deducted from the invoice price at time of sale, rather than billed back later."),
        ("Bill-back", "A funding mechanism where the retailer bills the manufacturer after the promotion executes, typically with documentation."),
        ("Scan-back", "A funding mechanism where the manufacturer pays based on actual units scanned during the promotion window."),
        ("Double-dip", "When a retailer collects the same promotional discount twice — once via off-invoice pricing and again via a billback deduction."),
        ("MCB", "Marketing contribution — billback. A retailer fee for marketing or merchandising support, billed after execution."),
        ("TPR", "Temporary price reduction. A short-term promotional discount on shelf price."),
        ("Ghost promo", "A promo_billback deduction referencing a promotion not found in the planned promotions calendar."),
        ("Deduction", "A dollar amount subtracted by the retailer from a remittance payment, with a reason code."),
        ("Dispute", "A formal challenge filed against a deduction, seeking full or partial recovery."),
        ("Recovery rate", "Percentage of disputed dollars successfully recovered (won + partial) out of total dollars disputed."),
    ]
    for term, definition in terms:
        row = _write_pair(ws, row, term, definition)

    # === SQL LOGIC ===
    row += 1
    row = _write_section(ws, row, "7. SQL Logic Summary")
    row = _write_body(ws, row,
        "Key queries that produce the locked numbers. All use cinderhaven_product_master.db.")
    row += 1

    sql_blocks = [
        ("Revenue (trailing 52w)",
         "SELECT SUM(dollars_sold) FROM scan_data\n"
         "WHERE week_ending >= (SELECT DISTINCT week_ending\n"
         "  FROM scan_data ORDER BY week_ending DESC LIMIT 1 OFFSET 51)"),
        ("Structural trade",
         "SELECT s.retailer, SUM(sd.dollars_sold) AS channel_rev\n"
         "FROM scan_data sd JOIN stores s ON sd.store_id = s.store_id\n"
         "WHERE sd.week_ending >= [trailing_52w_start]\n"
         "GROUP BY s.retailer\n"
         "-- Then: channel_rev × AVG(trade_spend_pct_[channel]) from sku_costs"),
        ("Operational waste",
         "SELECT SUM(amount) FROM deductions\n"
         "WHERE deduction_date > date([max_scan], '-365 days')\n"
         "  AND deduction_date <= [max_scan]\n"
         "  AND deduction_type != 'promo_billback'"),
        ("Recovery rate",
         "SELECT SUM(recovered_amount) / \n"
         "  (SELECT SUM(d.amount) FROM deductions d\n"
         "   JOIN disputes dis ON dis.deduction_id = d.deduction_id)\n"
         "FROM disputes"),
        ("Promo ROI",
         "-- Per promo: baseline = AVG(units) for N weeks pre-start\n"
         "-- incr_vol = (AVG(units during) - baseline) × duration_weeks\n"
         "-- incr_rev = incr_vol × AVG(dollars_sold/units_sold)\n"
         "-- ROI = incr_rev / COALESCE(actual_cost, planned_cost)"),
    ]
    for title, sql in sql_blocks:
        row = _write_pair(ws, row, title, sql)
