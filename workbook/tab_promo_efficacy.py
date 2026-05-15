"""Tab 3: Promo Efficacy — which promotions created value vs. destroyed it."""

from datetime import date

from workbook.db import connect

from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import bisect

from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_RIGHT,
    FILL_INPUT,
    FONT_BODY,
    FONT_HEADER,
    FONT_SECTION,
    FONT_SMALL,
    NUM_FMT_DOLLAR,
    NUM_FMT_PCT,
    TABLE_STYLE,
)

_HELPER_COL_START = 27
_MAX_WINDOW = 12


def _query_promo_data(database_url: str) -> dict:
    conn = connect()

    all_weeks = [r[0] for r in conn.execute(
        "SELECT DISTINCT week_ending FROM stg_scan_data ORDER BY week_ending"
    ).fetchall()]
    week_idx = {w: i for i, w in enumerate(all_weeks)}

    promos = conn.execute("""
        SELECT promo_id, sku, retailer, start_week, end_week,
               duration_weeks, discount_depth_pct, promo_type,
               promo_cost, funding_mechanism
        FROM stg_promotions
        ORDER BY promo_id, retailer
    """).fetchall()

    asp_map = {}
    asp_rows = conn.execute("""
        SELECT sd.sku, s.retailer, AVG(sd.dollars_sold * 1.0 / sd.units_sold)
        FROM stg_scan_data sd
        JOIN stg_stores s ON sd.store_id = s.store_id
        WHERE sd.units_sold > 0
        GROUP BY sd.sku, s.retailer
    """).fetchall()
    for sku, retailer, asp in asp_rows:
        asp_map[(sku, retailer)] = float(asp) if asp else None

    vol_rows = conn.execute("""
        SELECT sd.sku, s.retailer, sd.week_ending, SUM(sd.units_sold) as units
        FROM stg_scan_data sd
        JOIN stg_stores s ON sd.store_id = s.store_id
        GROUP BY sd.sku, s.retailer, sd.week_ending
        ORDER BY sd.sku, s.retailer, sd.week_ending
    """).fetchall()

    vol_map = {}
    for sku, retailer, week, units in vol_rows:
        vol_map.setdefault((sku, retailer), {})[week] = float(units) if units else 0

    matched_deductions = conn.execute("""
        SELECT p.promo_id, p.sku, p.retailer, SUM(d.amount) as actual_cost
        FROM stg_promotions p
        JOIN stg_deductions d ON LOWER(REPLACE(p.retailer, ' ', '_')) = d.retailer_id
            AND d.deduction_type = 'promo_billback'
            AND d.deduction_date BETWEEN (p.start_week - interval '14 days')::date
                                     AND (p.end_week + interval '90 days')::date
        GROUP BY p.promo_id, p.sku, p.retailer
    """).fetchall()
    actual_cost_map = {}
    for pid, sku, retailer, cost in matched_deductions:
        actual_cost_map[(pid, sku, retailer)] = float(cost) if cost else None

    ghosts = conn.execute("""
        SELECT d.deduction_id, d.retailer_id, d.amount, d.deduction_date
        FROM stg_deductions d
        WHERE d.deduction_type = 'promo_billback'
          AND NOT EXISTS (
              SELECT 1 FROM stg_promotions p
              WHERE LOWER(REPLACE(p.retailer, ' ', '_')) = d.retailer_id
                AND d.deduction_date BETWEEN (p.start_week - interval '14 days')::date
                                         AND (p.end_week + interval '90 days')::date
          )
        ORDER BY d.amount DESC
        LIMIT 20
    """).fetchall()

    ghost_summary = conn.execute("""
        SELECT COUNT(*), SUM(d.amount)
        FROM stg_deductions d
        WHERE d.deduction_type = 'promo_billback'
          AND NOT EXISTS (
              SELECT 1 FROM stg_promotions p
              WHERE LOWER(REPLACE(p.retailer, ' ', '_')) = d.retailer_id
                AND d.deduction_date BETWEEN (p.start_week - interval '14 days')::date
                                         AND (p.end_week + interval '90 days')::date
          )
    """).fetchone()

    conn.close()

    def _find_nearest_week_idx(target_date) -> int | None:
        idx = bisect.bisect_left(all_weeks, target_date)
        if idx >= len(all_weeks):
            return len(all_weeks) - 1 if all_weeks else None
        if idx == 0:
            return 0
        if abs((all_weeks[idx] - target_date).days) <= abs((all_weeks[idx - 1] - target_date).days):
            return idx
        return idx - 1

    promo_results = []
    for (pid, sku, retailer, start_wk, end_wk, dur_wks,
         discount, ptype, planned_cost, funding) in promos:

        weekly = vol_map.get((sku, retailer), {})
        asp = asp_map.get((sku, retailer))

        start_idx = _find_nearest_week_idx(start_wk)
        end_idx = _find_nearest_week_idx(end_wk)

        if start_idx is None or not weekly:
            promo_results.append({
                "promo_id": pid, "sku": sku, "retailer": retailer,
                "promo_type": ptype, "start_week": start_wk, "end_week": end_wk,
                "planned_cost": float(planned_cost) if planned_cost else None,
                "funding": funding,
                "actual_cost": actual_cost_map.get((pid, sku, retailer)),
                "asp": asp, "dur_wks": dur_wks,
                "pre_volumes": [], "during_volumes": [], "post_volumes": [],
                "data_quality": "No POS",
            })
            continue

        pre_volumes = []
        for offset in range(_MAX_WINDOW, 0, -1):
            idx = start_idx - offset
            if 0 <= idx < len(all_weeks):
                wk = all_weeks[idx]
                pre_volumes.append(weekly.get(wk, 0))
            else:
                pre_volumes.append(None)

        during_volumes = []
        for idx in range(start_idx, min(end_idx + 1, len(all_weeks))):
            wk = all_weeks[idx]
            during_volumes.append(weekly.get(wk, 0))

        post_volumes = []
        for offset in range(1, _MAX_WINDOW + 1):
            idx = end_idx + offset
            if idx < len(all_weeks):
                wk = all_weeks[idx]
                post_volumes.append(weekly.get(wk, 0))
            else:
                post_volumes.append(None)

        has_pre = any(v is not None and v > 0 for v in pre_volumes[:4])
        has_post = any(v is not None and v > 0 for v in post_volumes[:4])
        has_during = len(during_volumes) > 0 and any(v > 0 for v in during_volumes)

        if has_pre and has_during and has_post:
            quality = "Full"
        elif has_during and (has_pre or has_post):
            quality = "Partial"
        else:
            quality = "No POS"

        promo_results.append({
            "promo_id": pid, "sku": sku, "retailer": retailer,
            "promo_type": ptype, "start_week": start_wk, "end_week": end_wk,
            "planned_cost": float(planned_cost) if planned_cost else None,
            "funding": funding,
            "actual_cost": actual_cost_map.get((pid, sku, retailer)),
            "asp": asp, "dur_wks": dur_wks,
            "pre_volumes": pre_volumes, "during_volumes": during_volumes,
            "post_volumes": post_volumes, "data_quality": quality,
        })

    return {
        "promos": promo_results,
        "ghosts": ghosts,
        "ghost_count": ghost_summary[0],
        "ghost_total": float(ghost_summary[1] or 0),
        "total_promos": len(promos),
    }


def build_promo_efficacy(ws: Worksheet, database_url: str) -> None:
    data = _query_promo_data(database_url)

    ws.sheet_view.showGridLines = False

    visible_widths = {
        1: 3, 2: 14, 3: 16, 4: 12, 5: 12, 6: 12, 7: 12,
        8: 12, 9: 12, 10: 12, 11: 12, 12: 12, 13: 13, 14: 12, 15: 10, 16: 12, 17: 10,
    }
    for col, w in visible_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Header (rows 1-3) ---
    ws.merge_cells("B1:N1")
    ws["B1"] = "Promo Efficacy"
    ws["B1"].font = FONT_HEADER

    ws.merge_cells("B2:N2")
    ws["B2"] = (
        "ROI calculated as simple pre/during/post volume comparison. "
        "This is not a causal model — see Methodology tab for assumptions and limitations."
    )
    ws["B2"].font = FONT_SMALL

    ws.merge_cells("B3:N3")
    ws["B3"] = f"Built {date.today().isoformat()}"
    ws["B3"].font = FONT_SMALL

    # --- Adjustable window input (row 5) ---
    ws.cell(row=5, column=2, value="Pre/post comparison window (weeks):").font = FONT_BODY
    window_cell = ws.cell(row=5, column=4, value=4)
    window_cell.fill = FILL_INPUT
    window_cell.alignment = ALIGN_CENTER
    window_cell.protection = Protection(locked=False)
    window_cell.comment = Comment(
        "Number of weeks before and after the promotion used to calculate baseline volume. "
        "Changing this value recalculates all ROI figures below.",
        "System", width=280, height=80,
    )
    window_ref = "$D$5"

    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="12")
    dv.errorStyle = "stop"
    dv.showErrorMessage = True
    dv.error = "Enter an integer between 1 and 12"
    dv.errorTitle = "Invalid window"
    ws.add_data_validation(dv)
    dv.add(window_cell)

    # --- Coverage disclosure (row 7) ---
    full_ct = sum(1 for p in data["promos"] if p["data_quality"] == "Full")
    partial_ct = sum(1 for p in data["promos"] if p["data_quality"] == "Partial")
    no_pos_ct = sum(1 for p in data["promos"] if p["data_quality"] == "No POS")
    total_cost = sum(p["planned_cost"] or 0 for p in data["promos"])
    covered_cost = sum(p["planned_cost"] or 0 for p in data["promos"] if p["data_quality"] == "Full")

    ws.cell(row=7, column=2, value="Coverage Summary").font = FONT_SECTION
    ws.merge_cells("B8:N8")
    ws.cell(row=8, column=2, value=(
        f"Full POS data: {full_ct}/{data['total_promos']} promo-rows "
        f"({covered_cost/total_cost*100:.0f}% of planned spend covered)  |  "
        f"Partial: {partial_ct}  |  No POS: {no_pos_ct}"
    )).font = FONT_BODY

    # --- Performance table ---
    table_header_row = 10
    headers = [
        "Promo ID", "Retailer", "SKU", "Type", "Start", "End",
        "Planned $", "Actual $", "Pre Avg Vol", "During Avg Vol",
        "Post Avg Vol", "Incr. Volume", "Incr. Revenue", "ROI",
        "Funding", "Quality",
    ]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=table_header_row, column=c, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.alignment = ALIGN_CENTER

    ws.freeze_panes = f"B{table_header_row + 1}"

    def _sort_key(p):
        if p["data_quality"] == "No POS":
            return (2, 0)
        pre = p["pre_volumes"][-4:]
        pre_valid = [v for v in pre if v is not None and v > 0]
        if not pre_valid or not p.get("asp"):
            return (1, 0)
        baseline = sum(pre_valid) / len(pre_valid)
        during = p["during_volumes"]
        during_avg = sum(during) / len(during) if during else 0
        incr_vol = (during_avg - baseline) * (p["dur_wks"] or 1)
        cost = p["actual_cost"] or p["planned_cost"] or 1
        roi = (incr_vol * p["asp"]) / cost
        return (0, -roi)

    sorted_promos = sorted(data["promos"], key=_sort_key)

    pre_start_col = _HELPER_COL_START
    during_start_col = pre_start_col + _MAX_WINDOW
    post_start_col = during_start_col + _MAX_WINDOW

    for col in range(pre_start_col, post_start_col + _MAX_WINDOW):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    for i, promo in enumerate(sorted_promos):
        row = table_header_row + 1 + i

        ws.cell(row=row, column=2, value=promo["promo_id"])
        ws.cell(row=row, column=3, value=promo["retailer"])
        ws.cell(row=row, column=4, value=promo["sku"])
        ws.cell(row=row, column=5, value=promo["promo_type"])
        ws.cell(row=row, column=6, value=promo["start_week"])
        ws.cell(row=row, column=7, value=promo["end_week"])

        c_plan = ws.cell(row=row, column=8, value=promo["planned_cost"])
        c_plan.number_format = NUM_FMT_DOLLAR
        c_plan.alignment = ALIGN_RIGHT

        c_act = ws.cell(row=row, column=9, value=promo["actual_cost"])
        c_act.number_format = NUM_FMT_DOLLAR
        c_act.alignment = ALIGN_RIGHT

        for j, vol in enumerate(promo["pre_volumes"]):
            ws.cell(row=row, column=pre_start_col + j, value=vol if vol is not None else "")

        for j, vol in enumerate(promo["during_volumes"][:_MAX_WINDOW]):
            ws.cell(row=row, column=during_start_col + j, value=vol)

        for j, vol in enumerate(promo["post_volumes"]):
            ws.cell(row=row, column=post_start_col + j, value=vol if vol is not None else "")

        dur_wks = promo["dur_wks"] or 1
        asp = promo["asp"]

        if promo["data_quality"] == "No POS" or not asp:
            for col in range(10, 16):
                ws.cell(row=row, column=col, value=None)
        else:
            last_pre_col = get_column_letter(pre_start_col + _MAX_WINDOW - 1)
            pre_formula = f'=AVERAGE(OFFSET({last_pre_col}{row},0,1-{window_ref},1,{window_ref}))'
            c_pre = ws.cell(row=row, column=10, value=pre_formula)
            c_pre.number_format = '#,##0'
            c_pre.alignment = ALIGN_RIGHT

            during_range_start = get_column_letter(during_start_col)
            during_range_end = get_column_letter(during_start_col + min(dur_wks, _MAX_WINDOW) - 1)
            during_formula = f'=AVERAGE({during_range_start}{row}:{during_range_end}{row})'
            c_dur = ws.cell(row=row, column=11, value=during_formula)
            c_dur.number_format = '#,##0'
            c_dur.alignment = ALIGN_RIGHT

            first_post_col = get_column_letter(post_start_col)
            post_formula = f'=AVERAGE(OFFSET({first_post_col}{row},0,0,1,{window_ref}))'
            c_post = ws.cell(row=row, column=12, value=post_formula)
            c_post.number_format = '#,##0'
            c_post.alignment = ALIGN_RIGHT

            pre_ref = f"{get_column_letter(10)}{row}"
            dur_ref = f"{get_column_letter(11)}{row}"
            incr_vol_formula = f'=({dur_ref}-{pre_ref})*{dur_wks}'
            c_incr = ws.cell(row=row, column=13, value=incr_vol_formula)
            c_incr.number_format = '#,##0'
            c_incr.alignment = ALIGN_RIGHT

            incr_ref = f"{get_column_letter(13)}{row}"
            incr_rev_formula = f'={incr_ref}*{asp:.2f}'
            c_rev = ws.cell(row=row, column=14, value=incr_rev_formula)
            c_rev.number_format = NUM_FMT_DOLLAR
            c_rev.alignment = ALIGN_RIGHT

            rev_ref = f"{get_column_letter(14)}{row}"
            cost_ref = f"{get_column_letter(9)}{row}" if promo["actual_cost"] else f"{get_column_letter(8)}{row}"
            roi_formula = f'=IFERROR({rev_ref}/{cost_ref},"")'
            c_roi = ws.cell(row=row, column=15, value=roi_formula)
            c_roi.number_format = '0.0'
            c_roi.alignment = ALIGN_CENTER

        ws.cell(row=row, column=16, value=promo["funding"])

        c_qual = ws.cell(row=row, column=17, value=promo["data_quality"])
        c_qual.alignment = ALIGN_CENTER

    table_end_row = table_header_row + len(sorted_promos)

    # Excel Table
    table_ref = f"B{table_header_row}:Q{table_end_row}"
    promo_table = Table(displayName="tbl_PromoEfficacy", ref=table_ref)
    promo_table.tableStyleInfo = TABLE_STYLE
    ws.add_table(promo_table)

    # Conditional formatting on data quality (col Q = 17)
    qual_range = f"Q{table_header_row + 1}:Q{table_end_row}"
    ws.conditional_formatting.add(
        qual_range,
        CellIsRule(operator="equal", formula=['"Full"'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")),
    )
    ws.conditional_formatting.add(
        qual_range,
        CellIsRule(operator="equal", formula=['"Partial"'],
                   fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")),
    )
    ws.conditional_formatting.add(
        qual_range,
        CellIsRule(operator="equal", formula=['"No POS"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
    )

    # Conditional formatting on ROI (col O = 15): <1 red first, ≥1 green second
    roi_range = f"O{table_header_row + 1}:O{table_end_row}"
    ws.conditional_formatting.add(
        roi_range,
        CellIsRule(operator="lessThan", formula=["1"],
                   stopIfTrue=True,
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
    )
    ws.conditional_formatting.add(
        roi_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                   stopIfTrue=True,
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")),
    )

    # --- Ghost promos section ---
    ghost_row = table_end_row + 3
    ws.cell(row=ghost_row, column=2, value="Ghost Promos — Deductions Without Matching Promotion").font = FONT_SECTION

    ghost_row += 1
    ws.merge_cells(f"B{ghost_row}:G{ghost_row}")
    ws.cell(row=ghost_row, column=2, value=(
        f"{data['ghost_count']} promo_billback deductions (${data['ghost_total']:,.0f}) "
        f"have no matching promotion in the calendar. Top 20 shown below."
    )).font = FONT_BODY

    ghost_row += 1
    ghost_headers = ["Deduction ID", "Retailer", "Amount", "Date"]
    for c, h in enumerate(ghost_headers, 2):
        cell = ws.cell(row=ghost_row, column=c, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True)

    for i, (ded_id, retailer, amount, ded_date) in enumerate(data["ghosts"]):
        r = ghost_row + 1 + i
        ws.cell(row=r, column=2, value=ded_id)
        ws.cell(row=r, column=3, value=retailer.replace("_", " ").title())
        c_a = ws.cell(row=r, column=4, value=amount)
        c_a.number_format = NUM_FMT_DOLLAR
        c_a.alignment = ALIGN_RIGHT
        ws.cell(row=r, column=5, value=ded_date)
