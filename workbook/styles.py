"""Shared workbook styles — Lailara Design System v2."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import TableStyleInfo

SERIF = "Playfair Display"
SANS = "Source Sans 3"

# --- Lailara palette (complete) ---
CHICAGO_20 = "1f2e7a"
CHICAGO_10 = "141e52"
CHICAGO_70 = "8e9ad0"
CHICAGO_95 = "e8eaf4"
HK_35 = "158f75"
HK_70 = "6dcdb5"
HK_95 = "e4f5f0"
TOKYO_40 = "b82d4a"
TOKYO_70 = "e68a9a"
TOKYO_95 = "fbe9ed"
SINGAPORE_55 = "ee8a2a"
SINGAPORE_70 = "f6b97c"
SINGAPORE_95 = "fdeee0"
RED_42 = "cc100a"
LONDON_5 = "0d0d0d"
LONDON_20 = "333333"
LONDON_35 = "595959"
LONDON_70 = "b3b3b3"
LONDON_85 = "d9d9d9"

# --- Fonts (Playfair Display for editorial, Source Sans 3 for functional) ---
FONT_HEADER = Font(name=SERIF, size=16, bold=True, color=LONDON_5)
FONT_SECTION = Font(name=SERIF, size=13, bold=True, color=LONDON_5)
FONT_KPI_VALUE = Font(name=SERIF, size=20, bold=True, color=CHICAGO_20)
FONT_KPI_LABEL = Font(name=SANS, size=11, color=LONDON_35)
FONT_BODY = Font(name=SANS, size=11, color=LONDON_20)
FONT_SMALL = Font(name=SANS, size=10, italic=True, color=LONDON_35)
FONT_NAV = Font(name=SANS, size=10, underline="single", color=CHICAGO_20)
FONT_TABLE_HEADER = Font(name=SANS, size=11, bold=True, color="FFFFFF")

# --- Fills ---
FILL_HEADER = PatternFill(start_color=CHICAGO_20, end_color=CHICAGO_20, fill_type="solid")
FILL_INPUT = PatternFill(start_color=SINGAPORE_95, end_color=SINGAPORE_95, fill_type="solid")
FILL_GOOD = PatternFill(start_color=HK_95, end_color=HK_95, fill_type="solid")
FILL_WARN = PatternFill(start_color=SINGAPORE_95, end_color=SINGAPORE_95, fill_type="solid")
FILL_BAD = PatternFill(start_color=TOKYO_95, end_color=TOKYO_95, fill_type="solid")

# --- Alignment ---
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# --- Borders ---
BORDER_SECTION = Border(bottom=Side(style="thin", color=LONDON_85))

# --- Number formats ---
NUM_FMT_DOLLAR = '#,##0'
NUM_FMT_PCT = '0.0%'

# --- Table style ---
TABLE_STYLE = TableStyleInfo(
    name="TableStyleLight1", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False,
)

# --- Tab names (canonical order, used by generator and nav links) ---
TAB_NAMES = (
    "Executive Pulse", "Leak Diagnostic", "Promo Efficacy",
    "Retailer Risk", "Deduction Ledger", "Deduction Code Crosswalk",
    "Methodology & Logic",
)
