"""Tab 5: Deduction Ledger — full trailing-365 deduction log for investigation."""

import contextlib
import sqlite3
from datetime import date
from pathlib import Path

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from workbook.queries import get_trailing_bounds
from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_RIGHT,
    FONT_HEADER,
    FONT_SMALL,
    NUM_FMT_DOLLAR,
    SANS,
    TABLE_STYLE,
)

COLUMNS = [
    ("Deduction ID", 14),
    ("Date", 11),
    ("Retailer", 16),
    ("Raw Code", 10),
    ("Translated Code", 30),
    ("Category", 14),
    ("Amount", 12),
    ("Order Ref", 12),
    ("Shipment Ref", 12),
    ("Remittance ID", 14),
    ("Remittance Desc", 24),
    ("Dispute Status", 14),
    ("Recovered", 11),
    ("Dispute Filed", 11),
    ("Dispute Closed", 12),
    ("Days Outstanding", 10),
    ("Deadline", 11),
    ("Vague", 6),
    ("Post-Audit", 9),
    ("Double-Dip", 9),
]


def _query_ledger(db_path: Path) -> tuple[list[tuple], str, str]:
    with contextlib.closing(sqlite3.connect(db_path)) as conn:
        oldest_week, max_scan = get_trailing_bounds(conn)

        rows = conn.execute("""
            SELECT
                d.deduction_id,
                d.deduction_date,
                d.retailer_id,
                d.code_as_remitted,
                COALESCE(dc.name, 'Unmapped'),
                d.deduction_type,
                d.amount,
                d.order_id,
                d.shipment_id,
                d.remittance_id,
                d.remittance_description,
                dis.outcome,
                dis.recovered_amount,
                dis.filed_date,
                dis.closed_date,
                CASE
                    WHEN dis.closed_date IS NOT NULL
                    THEN CAST(julianday(dis.closed_date) - julianday(d.deduction_date) AS INTEGER)
                    WHEN dis.filed_date IS NOT NULL
                    THEN CAST(julianday(?) - julianday(d.deduction_date) AS INTEGER)
                    ELSE NULL
                END,
                d.dispute_deadline,
                d.is_vague,
                d.is_post_audit,
                d.is_double_dip
            FROM deductions d
            LEFT JOIN deduction_codes dc ON d.code_id = dc.code_id
            LEFT JOIN (
                SELECT deduction_id,
                       MAX(outcome) AS outcome,
                       SUM(recovered_amount) AS recovered_amount,
                       MIN(filed_date) AS filed_date,
                       MAX(closed_date) AS closed_date
                FROM disputes
                GROUP BY deduction_id
            ) dis ON dis.deduction_id = d.deduction_id
            WHERE d.deduction_date > date(?, '-365 days') AND d.deduction_date <= ?
            ORDER BY d.deduction_date DESC, d.amount DESC
        """, (max_scan, max_scan, max_scan)).fetchall()

    return rows, oldest_week, max_scan


def build_deduction_ledger(ws: Worksheet, db_path: Path) -> None:
    rows, oldest_week, max_scan = _query_ledger(db_path)

    ws.sheet_view.showGridLines = True

    # --- Header ---
    ws.merge_cells("A1:F1")
    ws["A1"] = "Deduction Ledger"
    ws["A1"].font = FONT_HEADER

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"{len(rows):,} deductions  |  "
        f"Trailing 365 days ({oldest_week} to {max_scan})  |  "
        f"Built {date.today().isoformat()}"
    )
    ws["A2"].font = FONT_SMALL

    # --- Column headers (row 4) ---
    header_row = 4
    header_font = Font(name=SANS, size=10, bold=True)

    for c, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.font = header_font
        cell.alignment = ALIGN_CENTER
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.freeze_panes = "B5"

    # --- Data rows ---
    for i, row_data in enumerate(rows):
        rw = header_row + 1 + i

        for c, val in enumerate(row_data, 1):
            ws.cell(row=rw, column=c, value=val)

        ws.cell(row=rw, column=7).number_format = NUM_FMT_DOLLAR
        ws.cell(row=rw, column=7).alignment = ALIGN_RIGHT
        ws.cell(row=rw, column=13).number_format = NUM_FMT_DOLLAR
        ws.cell(row=rw, column=13).alignment = ALIGN_RIGHT
        ws.cell(row=rw, column=16).alignment = ALIGN_CENTER
        for flag_col in (18, 19, 20):
            cell = ws.cell(row=rw, column=flag_col)
            cell.value = "Yes" if cell.value == 1 else ""
            cell.alignment = ALIGN_CENTER

    # --- Excel Table ---
    last_col = get_column_letter(len(COLUMNS))
    table_end = header_row + len(rows)
    table_ref = f"A{header_row}:{last_col}{table_end}"

    table = Table(displayName="tbl_DeductionLedger", ref=table_ref)
    table.tableStyleInfo = TABLE_STYLE
    ws.add_table(table)
