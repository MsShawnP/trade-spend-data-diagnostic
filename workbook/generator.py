"""Generate the trade spend diagnostic workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from workbook.tab_code_crosswalk import build_code_crosswalk
from workbook.tab_deduction_ledger import build_deduction_ledger
from workbook.tab_executive_pulse import build_executive_pulse
from workbook.tab_leak_diagnostic import build_leak_diagnostic
from workbook.tab_methodology import build_methodology
from workbook.tab_promo_efficacy import build_promo_efficacy
from workbook.tab_retailer_risk import build_retailer_risk


TAB_SPEC = [
    ("Executive Pulse", "00B050"),
    ("Leak Diagnostic", "00B050"),
    ("Promo Efficacy", "00B050"),
    ("Retailer Risk", "00B050"),
    ("Deduction Ledger", "4472C4"),
    ("Deduction Code Crosswalk", "A5A5A5"),
    ("Methodology & Logic", "A5A5A5"),
]


def _add_named_ranges(wb: Workbook) -> None:
    """Define named ranges for key metrics referenced by other tools."""
    ranges = {
        "AllInTradeRate": "'Executive Pulse'!$B$5",
        "StructuralTradeRate": "'Executive Pulse'!$C$5",
        "OperationalWasteRate": "'Executive Pulse'!$D$5",
        "TotalRevenue": "'Executive Pulse'!$D$11",
        "StructuralTrade": "'Executive Pulse'!$D$12",
        "OperationalWaste": "'Executive Pulse'!$D$13",
        "AllInTradeCost": "'Executive Pulse'!$D$12+'Executive Pulse'!$D$13",
        "RecoveryRate": "'Executive Pulse'!$C$25",
        "KPI_AllInTradeRate": "'Executive Pulse'!$B$5",
        "KPI_PlannedTradeRate": "'Executive Pulse'!$C$5",
        "KPI_OperationalWaste": "'Executive Pulse'!$D$5",
        "KPI_Revenue": "'Executive Pulse'!$D$11",
        "KPI_StructuralTrade": "'Executive Pulse'!$D$12",
        "KPI_OpWasteAmount": "'Executive Pulse'!$D$13",
    }
    for name, ref in ranges.items():
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names.add(dn)


def _set_print_areas(wb: Workbook) -> None:
    """Set print areas for analysis tabs (landscape letter)."""
    wb["Executive Pulse"].print_area = "A1:F58"
    wb["Leak Diagnostic"].print_area = "A1:G55"
    wb["Promo Efficacy"].print_area = "A1:Q9"
    wb["Retailer Risk"].print_area = "A1:M17"

    for tab_name in ["Executive Pulse", "Leak Diagnostic", "Promo Efficacy", "Retailer Risk"]:
        ws = wb[tab_name]
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER


def generate_workbook(db_path: Path, output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    for name, color in TAB_SPEC:
        ws = wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = color

    build_executive_pulse(wb["Executive Pulse"], db_path)
    build_leak_diagnostic(wb["Leak Diagnostic"], db_path)
    build_promo_efficacy(wb["Promo Efficacy"], db_path)
    build_retailer_risk(wb["Retailer Risk"], db_path)
    build_deduction_ledger(wb["Deduction Ledger"], db_path)
    build_code_crosswalk(wb["Deduction Code Crosswalk"], db_path)
    build_methodology(wb["Methodology & Logic"])

    _add_named_ranges(wb)
    _set_print_areas(wb)

    # Set active sheet to Tab 1
    wb.active = wb.sheetnames.index("Executive Pulse")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
