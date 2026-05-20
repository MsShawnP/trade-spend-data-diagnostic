"""Tab 4: Retailer Risk — net-net margin by retailer with what-if scenarios."""

import contextlib
import sqlite3
from datetime import date
from pathlib import Path

from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from workbook.queries import (
    CHANNEL_RATE_COLS,
    fetch_channel_rates,
    get_trailing_bounds,
    retailer_key,
)
from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_RIGHT,
    FILL_INPUT,
    FONT_BODY,
    FONT_HEADER,
    FONT_SECTION,
    FONT_SMALL,
    HK_35,
    NUM_FMT_DOLLAR,
    NUM_FMT_PCT,
    SANS,
    SINGAPORE_55,
    TABLE_STYLE,
    TOKYO_40,
)

REGIONAL_RETAILERS = [
    "Kroger", "Sprouts", "Regional Group",
]


def _query_retailer_data(db_path: Path) -> dict:
    with contextlib.closing(sqlite3.connect(db_path)) as conn:
        oldest_week, max_scan = get_trailing_bounds(conn)

        rev_rows = conn.execute("""
            SELECT s.retailer, SUM(sd.dollars_sold)
            FROM scan_data sd JOIN stores s ON sd.store_id = s.store_id
            WHERE sd.week_ending >= ?
            GROUP BY s.retailer
        """, (oldest_week,)).fetchall()
        revenue_map = dict(rev_rows)
        total_revenue = sum(revenue_map.values())

        rates, regional_rate = fetch_channel_rates(conn)

        gm_map: dict[str, float] = {}
        for channel, wcol in [("Walmart", "wholesale_walmart"), ("Costco", "wholesale_costco"),
                              ("Whole Foods", "wholesale_whole_foods"), ("UNFI", "wholesale_unfi"),
                              ("DTC", "wholesale_dtc"), ("KeHE", "wholesale_kehe"),
                              ("Regional", "wholesale_regional")]:
            r = conn.execute(f"SELECT AVG(cogs_per_unit), AVG({wcol}) FROM sku_costs").fetchone()
            if r[0] is not None and r[1]:
                gm_map[channel] = (r[1] - r[0]) / r[1]
            else:
                gm_map[channel] = 0

        op_ded_rows = conn.execute("""
            SELECT retailer_id, SUM(amount)
            FROM deductions
            WHERE deduction_date > date(?, '-365 days') AND deduction_date <= ?
              AND deduction_type != 'promo_billback'
            GROUP BY retailer_id
        """, (max_scan, max_scan)).fetchall()
        op_ded_map = dict(op_ded_rows)

        pb_rows = conn.execute("""
            SELECT retailer_id, SUM(amount)
            FROM deductions
            WHERE deduction_date > date(?, '-365 days') AND deduction_date <= ?
              AND deduction_type = 'promo_billback'
            GROUP BY retailer_id
        """, (max_scan, max_scan)).fetchall()
        pb_map = dict(pb_rows)

    channel_order = ["Walmart", "UNFI", "KeHE", "Whole Foods", "Costco", "DTC"] + REGIONAL_RETAILERS

    retailers = []
    for retailer_name in channel_order:
        rev = revenue_map.get(retailer_name, 0)
        rate = rates.get(retailer_name, regional_rate)
        structural = rev * rate

        rkey = retailer_key(retailer_name)
        op_ded = op_ded_map.get(rkey, 0)
        pb_ded = pb_map.get(rkey, 0)

        all_in = structural + op_ded + pb_ded
        all_in_rate = all_in / rev if rev else 0

        gm_key = retailer_name if retailer_name in gm_map else "Regional"
        gross_margin = gm_map.get(gm_key, gm_map.get("Regional", 0.4))

        total_ded = op_ded + pb_ded

        retailers.append({
            "name": retailer_name,
            "revenue": rev,
            "rev_share": rev / total_revenue if total_revenue else 0,
            "structural": structural,
            "structural_rate": rate,
            "op_deductions": op_ded,
            "op_ded_rate": op_ded / rev if rev else 0,
            "pb_deductions": pb_ded,
            "all_in": all_in,
            "all_in_rate": all_in_rate,
            "gross_margin": gross_margin,
            "net_net_margin": gross_margin - (all_in / rev if rev else 0),
            "total_deductions": total_ded,
        })

    total_deductions = sum(r["total_deductions"] for r in retailers)
    for r in retailers:
        r["ded_share"] = r["total_deductions"] / total_deductions if total_deductions else 0

    return {
        "retailers": retailers,
        "total_revenue": total_revenue,
        "oldest_week": oldest_week,
        "max_scan": max_scan,
    }


def build_retailer_risk(ws: Worksheet, db_path: Path) -> None:
    data = _query_retailer_data(db_path)
    retailers = data["retailers"]

    ws.sheet_view.showGridLines = False

    col_widths = [3, 20, 14, 10, 14, 10, 14, 10, 13, 14, 10, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Header ---
    ws.merge_cells("B1:L1")
    ws["B1"] = "Retailer Risk"
    ws["B1"].font = FONT_HEADER

    ws.merge_cells("B2:L2")
    ws["B2"] = f"Trailing 52 weeks ({data['oldest_week']} to {data['max_scan']})  |  Built {date.today().isoformat()}"
    ws["B2"].font = FONT_SMALL

    # --- P&L summary table ---
    row = 4
    ws.cell(row=row, column=2, value="Retailer P&L Summary").font = FONT_SECTION

    row += 1
    headers = [
        "Retailer", "Revenue", "Rev %", "Structural $",
        "Struct %", "Op Deductions", "Op Ded %", "Promo BB",
        "All-In Trade", "All-In %", "Gross Margin", "Net-Net Margin",
    ]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name=SANS, size=10, bold=True)
        cell.alignment = ALIGN_CENTER

    table_start = row + 1
    for i, r in enumerate(retailers):
        rw = table_start + i

        ws.cell(row=rw, column=2, value=r["name"])

        c = ws.cell(row=rw, column=3, value=r["revenue"])
        c.number_format = NUM_FMT_DOLLAR
        c.alignment = ALIGN_RIGHT

        c = ws.cell(row=rw, column=4, value=r["rev_share"])
        c.number_format = NUM_FMT_PCT
        c.alignment = ALIGN_CENTER

        c = ws.cell(row=rw, column=5, value=r["structural"])
        c.number_format = NUM_FMT_DOLLAR
        c.alignment = ALIGN_RIGHT

        c = ws.cell(row=rw, column=6, value=r["structural_rate"])
        c.number_format = NUM_FMT_PCT
        c.alignment = ALIGN_CENTER

        c = ws.cell(row=rw, column=7, value=r["op_deductions"])
        c.number_format = NUM_FMT_DOLLAR
        c.alignment = ALIGN_RIGHT

        c = ws.cell(row=rw, column=8, value=r["op_ded_rate"])
        c.number_format = NUM_FMT_PCT
        c.alignment = ALIGN_CENTER

        c = ws.cell(row=rw, column=9, value=r["pb_deductions"])
        c.number_format = NUM_FMT_DOLLAR
        c.alignment = ALIGN_RIGHT

        c = ws.cell(row=rw, column=10, value=r["all_in"])
        c.number_format = NUM_FMT_DOLLAR
        c.alignment = ALIGN_RIGHT

        c = ws.cell(row=rw, column=11, value=r["all_in_rate"])
        c.number_format = NUM_FMT_PCT
        c.alignment = ALIGN_CENTER

        c = ws.cell(row=rw, column=12, value=r["gross_margin"])
        c.number_format = NUM_FMT_PCT
        c.alignment = ALIGN_CENTER

        c = ws.cell(row=rw, column=13, value=r["net_net_margin"])
        c.number_format = NUM_FMT_PCT
        c.alignment = ALIGN_CENTER

    table_end = table_start + len(retailers) - 1

    pnl_table = Table(displayName="tbl_RetailerPnL", ref=f"B{row}:M{table_end}")
    pnl_table.tableStyleInfo = TABLE_STYLE
    ws.add_table(pnl_table)

    margin_range = f"M{table_start}:M{table_end}"
    ws.conditional_formatting.add(
        margin_range,
        ColorScaleRule(
            start_type="num", start_value=0, start_color=TOKYO_40,
            mid_type="num", mid_value=0.25, mid_color=SINGAPORE_55,
            end_type="num", end_value=0.5, end_color=HK_35,
        ),
    )

    # --- Concentration risk data table ---
    chart_retailers = [r for r in retailers if r["revenue"] > 0]

    cr_section_row = table_end + 3
    ws.cell(row=cr_section_row, column=2, value="Concentration Risk: Revenue Share vs. Deduction Share").font = FONT_SECTION

    cr_header_row = cr_section_row + 1
    ws.cell(row=cr_header_row, column=2, value="Retailer").font = FONT_BODY
    ws.cell(row=cr_header_row, column=3, value="Revenue Share").font = FONT_BODY
    ws.cell(row=cr_header_row, column=4, value="Deduction Share").font = FONT_BODY

    for i, r in enumerate(chart_retailers):
        rw = cr_header_row + 1 + i
        ws.cell(row=rw, column=2, value=r["name"])
        ws.cell(row=rw, column=3, value=r["rev_share"]).number_format = NUM_FMT_PCT
        ws.cell(row=rw, column=4, value=r["ded_share"]).number_format = NUM_FMT_PCT

    cr_end_row = cr_header_row + len(chart_retailers)

    cr_table = Table(displayName="tbl_ConcentrationRisk", ref=f"B{cr_header_row}:D{cr_end_row}")
    cr_table.tableStyleInfo = TABLE_STYLE
    ws.add_table(cr_table)

    if len(chart_retailers) < len(retailers):
        note_row = cr_end_row + 1
        ws.cell(row=note_row, column=2,
                value=f"Note: {len(retailers) - len(chart_retailers)} retailer(s) with zero revenue excluded from scenario tables."
                ).font = FONT_SMALL

    # --- What-if trade rate inputs ---
    whatif_row = cr_end_row + 2
    ws.cell(row=whatif_row, column=2, value="What-If Trade Rate Scenario").font = FONT_SECTION

    whatif_row += 1
    ws.merge_cells(f"B{whatif_row}:I{whatif_row}")
    ws.cell(row=whatif_row, column=2,
            value="Enter target all-in trade rates below. Savings show annual impact of achieving target."
            ).font = FONT_SMALL

    whatif_row += 1
    whatif_headers = ["Retailer", "Revenue", "Current Rate", "Target Rate",
                      "Trade at Target", "Current Trade", "Annual Savings", "What-If Margin"]
    for c, h in enumerate(whatif_headers, 2):
        cell = ws.cell(row=whatif_row, column=c, value=h)
        cell.font = Font(name=SANS, size=10, bold=True)
        cell.alignment = ALIGN_CENTER

    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="0.5")
    dv.error = "Enter a rate between 0% and 50%"
    dv.errorTitle = "Invalid rate"
    ws.add_data_validation(dv)

    for i, r in enumerate(chart_retailers):
        rw = whatif_row + 1 + i

        ws.cell(row=rw, column=2, value=r["name"])

        rev_cell = ws.cell(row=rw, column=3, value=r["revenue"])
        rev_cell.number_format = NUM_FMT_DOLLAR
        rev_cell.alignment = ALIGN_RIGHT

        cur_cell = ws.cell(row=rw, column=4, value=r["all_in_rate"])
        cur_cell.number_format = NUM_FMT_PCT
        cur_cell.alignment = ALIGN_CENTER

        target_cell = ws.cell(row=rw, column=5, value=r["all_in_rate"])
        target_cell.number_format = NUM_FMT_PCT
        target_cell.fill = FILL_INPUT
        target_cell.alignment = ALIGN_CENTER
        target_cell.protection = Protection(locked=False)
        target_cell.comment = Comment(
            "Enter a target all-in trade rate for this retailer. "
            "The savings and margin columns show the annual impact.",
            "System", width=260, height=60,
        )
        dv.add(target_cell)

        ws.cell(row=rw, column=6, value=f"=C{rw}*E{rw}").number_format = NUM_FMT_DOLLAR
        ws.cell(row=rw, column=6).alignment = ALIGN_RIGHT

        cur_trade = ws.cell(row=rw, column=7, value=r["all_in"])
        cur_trade.number_format = NUM_FMT_DOLLAR
        cur_trade.alignment = ALIGN_RIGHT

        ws.cell(row=rw, column=8, value=f"=G{rw}-F{rw}").number_format = NUM_FMT_DOLLAR
        ws.cell(row=rw, column=8).alignment = ALIGN_RIGHT

        gm_literal = f"{r['gross_margin']:.6f}"
        ws.cell(row=rw, column=9, value=f"={gm_literal}-E{rw}").number_format = NUM_FMT_PCT
        ws.cell(row=rw, column=9).alignment = ALIGN_CENTER
