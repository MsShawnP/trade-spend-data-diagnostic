"""Tab 2: Leak Diagnostic — where the operational waste is and what's recoverable."""

import sqlite3
from datetime import date
from pathlib import Path

from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from workbook.styles import (
    ALIGN_CENTER,
    ALIGN_RIGHT,
    FILL_BAD,
    FILL_GOOD,
    FILL_INPUT,
    FILL_WARN,
    FONT_BODY,
    FONT_HEADER,
    FONT_SECTION,
    FONT_SMALL,
    NUM_FMT_DOLLAR,
    NUM_FMT_PCT,
    SANS,
)

RECOVERABILITY = {
    "vague": "Low",
    "label_fine": "Low",
    "short_ship": "Medium",
    "spoilage": "Low",
    "late_delivery": "Medium",
    "slotting": "Low",
    "damaged": "Medium",
    "pallet_fine": "Low",
}

TABLE_STYLE = TableStyleInfo(
    name="TableStyleLight1", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False,
)


def _query_data(db_path: Path) -> dict:
    conn = sqlite3.connect(db_path)

    weeks = conn.execute(
        "SELECT DISTINCT week_ending FROM scan_data ORDER BY week_ending DESC LIMIT 52"
    ).fetchall()
    oldest_week = weeks[-1][0]
    max_scan = weeks[0][0]

    categories = conn.execute("""
        SELECT
            d.deduction_type,
            COUNT(*) as cnt,
            SUM(d.amount) as total,
            AVG(
                CASE WHEN dis.closed_date IS NOT NULL
                     THEN julianday(dis.closed_date) - julianday(d.deduction_date)
                END
            ) as avg_days
        FROM deductions d
        LEFT JOIN disputes dis ON dis.deduction_id = d.deduction_id
        WHERE d.deduction_date > date(?, '-365 days') AND d.deduction_date <= ?
          AND d.deduction_type != 'promo_billback'
        GROUP BY d.deduction_type
        ORDER BY SUM(d.amount) DESC
    """, (max_scan, max_scan)).fetchall()

    operational_waste = sum(r[2] for r in categories)

    double_dips = conn.execute("""
        SELECT deduction_id, retailer_id, amount, deduction_date, deduction_type
        FROM deductions
        WHERE is_double_dip = 1
        ORDER BY amount DESC
    """).fetchall()

    recovery = conn.execute("""
        SELECT COUNT(*), SUM(recovered_amount)
        FROM disputes
    """).fetchone()

    conn.close()

    return {
        "categories": categories,
        "operational_waste": operational_waste,
        "double_dips": double_dips,
        "dispute_count": recovery[0],
        "total_recovered": recovery[1],
        "oldest_week": oldest_week,
        "max_scan": max_scan,
    }


def build_leak_diagnostic(ws: Worksheet, db_path: Path) -> None:
    data = _query_data(db_path)

    ws.sheet_view.showGridLines = False

    col_widths = [3, 22, 14, 16, 16, 20, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Header (rows 1-3) ---
    ws.merge_cells("B1:F1")
    ws["B1"] = "Leak Diagnostic"
    ws["B1"].font = FONT_HEADER

    ws.merge_cells("B2:F2")
    ws["B2"] = f"Trailing 365 days ({data['oldest_week']} to {data['max_scan']})  |  Built {date.today().isoformat()}"
    ws["B2"].font = FONT_SMALL

    # --- Category breakdown table (row 4+) ---
    row = 4
    ws.cell(row=row, column=2, value="Operational Waste by Category").font = FONT_SECTION

    row += 1
    headers = ["Category", "Count", "Total Amount", "% of Waste", "Avg Days to Resolve", "Recoverability"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name=SANS, size=11, bold=True)
        cell.alignment = ALIGN_CENTER

    table_start_row = row + 1
    for i, (dtype, cnt, total, avg_days) in enumerate(data["categories"]):
        r = table_start_row + i
        label = dtype.replace("_", " ").title()
        pct = total / data["operational_waste"]
        recov = RECOVERABILITY.get(dtype, "Low")

        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=cnt).alignment = ALIGN_RIGHT
        c_amt = ws.cell(row=r, column=4, value=total)
        c_amt.number_format = NUM_FMT_DOLLAR
        c_amt.alignment = ALIGN_RIGHT
        c_pct = ws.cell(row=r, column=5, value=pct)
        c_pct.number_format = NUM_FMT_PCT
        c_pct.alignment = ALIGN_CENTER
        c_days = ws.cell(row=r, column=6, value=round(avg_days) if avg_days else None)
        c_days.alignment = ALIGN_CENTER
        ws.cell(row=r, column=7, value=recov).alignment = ALIGN_CENTER

    table_end_row = table_start_row + len(data["categories"]) - 1

    # Excel Table for category breakdown
    cat_table = Table(displayName="tbl_WasteByCategory", ref=f"B{row}:G{table_end_row}")
    cat_table.tableStyleInfo = TABLE_STYLE
    ws.add_table(cat_table)

    # Conditional formatting on recoverability column
    recov_range = f"G{table_start_row}:G{table_end_row}"
    ws.conditional_formatting.add(
        recov_range,
        CellIsRule(operator="equal", formula=['"High"'], fill=FILL_GOOD),
    )
    ws.conditional_formatting.add(
        recov_range,
        CellIsRule(operator="equal", formula=['"Medium"'], fill=FILL_WARN),
    )
    ws.conditional_formatting.add(
        recov_range,
        CellIsRule(operator="equal", formula=['"Low"'], fill=FILL_BAD),
    )

    # Totals row (outside table)
    totals_row = table_end_row + 1
    ws.cell(row=totals_row, column=2, value="Total").font = Font(name=SANS, size=11, bold=True)
    c_tcnt = ws.cell(row=totals_row, column=3, value=sum(r[1] for r in data["categories"]))
    c_tcnt.alignment = ALIGN_RIGHT
    c_tcnt.font = Font(name=SANS, size=11, bold=True)
    c_tamt = ws.cell(row=totals_row, column=4, value=data["operational_waste"])
    c_tamt.number_format = NUM_FMT_DOLLAR
    c_tamt.alignment = ALIGN_RIGHT
    c_tamt.font = Font(name=SANS, size=11, bold=True)

    # --- Double-dip alert (compact — no chart gap) ---
    dd_row = totals_row + 2
    ws.cell(row=dd_row, column=2, value="Double-Dip Alert").font = FONT_SECTION

    dd_row += 1
    ws.merge_cells(f"B{dd_row}:F{dd_row}")
    ws.cell(row=dd_row, column=2,
            value=f"{len(data['double_dips'])} events detected — ${sum(d[2] for d in data['double_dips']):,.0f} total double-payment"
            ).font = FONT_BODY

    dd_row += 1
    dd_header_row = dd_row
    dd_headers = ["Deduction ID", "Retailer", "Amount", "Date", "Type"]
    for c, h in enumerate(dd_headers, 2):
        cell = ws.cell(row=dd_row, column=c, value=h)
        cell.font = Font(name=SANS, size=11, bold=True)

    for i, (ded_id, retailer, amount, ded_date, dtype) in enumerate(data["double_dips"]):
        r = dd_row + 1 + i
        ws.cell(row=r, column=2, value=ded_id).comment = Comment(
            "This deduction represents a double-payment: the promotion received both "
            "an off-invoice discount on the original invoice and a promo_billback "
            "deduction, resulting in Cinderhaven paying twice for the same promotion.",
            "System", width=300, height=100,
        )
        ws.cell(row=r, column=3, value=retailer.replace("_", " ").title())
        c_a = ws.cell(row=r, column=4, value=amount)
        c_a.number_format = NUM_FMT_DOLLAR
        c_a.alignment = ALIGN_RIGHT
        ws.cell(row=r, column=5, value=ded_date)
        ws.cell(row=r, column=6, value=dtype.replace("_", " ").title())

    dd_end_row = dd_row + len(data["double_dips"])

    # Excel Table for double-dips
    if data["double_dips"]:
        dd_table = Table(displayName="tbl_DoubleDips", ref=f"B{dd_header_row}:F{dd_end_row}")
        dd_table.tableStyleInfo = TABLE_STYLE
        ws.add_table(dd_table)

    # --- Adjustable target recovery rate ---
    recov_row = dd_end_row + 2
    ws.cell(row=recov_row, column=2, value="Recovery Opportunity Model").font = FONT_SECTION

    recov_row += 1
    ws.cell(row=recov_row, column=2, value="Target recovery rate:").font = FONT_BODY
    input_cell = ws.cell(row=recov_row, column=3, value=0.30)
    input_cell.number_format = NUM_FMT_PCT
    input_cell.fill = FILL_INPUT
    input_cell.alignment = ALIGN_CENTER
    input_cell.protection = Protection(locked=False)
    input_cell.comment = Comment(
        "Adjustable input: enter your target recovery rate (0%–100%). "
        "Cells below will recalculate automatically.",
        "System", width=250, height=60,
    )
    input_ref = f"C{recov_row}"

    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
    dv.error = "Please enter a value between 0% and 100%"
    dv.errorTitle = "Invalid recovery rate"
    ws.add_data_validation(dv)
    dv.add(input_cell)

    recov_row += 1
    waste_cell_row = recov_row
    ws.cell(row=recov_row, column=2, value="Operational waste (base):").font = FONT_BODY
    ws.cell(row=recov_row, column=3, value=data["operational_waste"]).number_format = NUM_FMT_DOLLAR

    recov_row += 1
    ws.cell(row=recov_row, column=2, value="Current recovery (19.8%):").font = FONT_BODY
    ws.cell(row=recov_row, column=3, value=data["total_recovered"]).number_format = NUM_FMT_DOLLAR

    recov_row += 1
    ws.cell(row=recov_row, column=2, value="Recovery at target rate:").font = FONT_BODY
    ws.cell(row=recov_row, column=3, value=f"=C{waste_cell_row}*{input_ref}").number_format = NUM_FMT_DOLLAR

    recov_row += 1
    ws.cell(row=recov_row, column=2, value="Incremental opportunity:").font = FONT_BODY
    ws.cell(row=recov_row, column=3, value=f"=C{recov_row-1}-C{recov_row-2}").number_format = NUM_FMT_DOLLAR
