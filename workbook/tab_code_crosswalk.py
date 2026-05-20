"""Tab 6: Deduction Code Crosswalk — reference mapping of retailer codes."""

import contextlib
import sqlite3
from pathlib import Path

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from workbook.styles import (
    ALIGN_CENTER,
    FILL_GOOD,
    FILL_WARN,
    FONT_HEADER,
    FONT_SMALL,
    SANS,
    TABLE_STYLE,
)

COLUMNS = [
    ("Retailer", 20),
    ("Retailer Code", 13),
    ("Description", 36),
    ("Category", 16),
    ("Status", 10),
]


def _query_crosswalk(db_path: Path) -> list[tuple]:
    with contextlib.closing(sqlite3.connect(db_path)) as conn:
        rows = conn.execute("""
            SELECT
                retailer_id,
                code,
                name,
                deduction_type,
                CASE WHEN is_published = 1 THEN 'Verified' ELSE 'Inferred' END
            FROM deduction_codes
            ORDER BY retailer_id, deduction_type, code
        """).fetchall()
    return rows


def build_code_crosswalk(ws: Worksheet, db_path: Path) -> None:
    rows = _query_crosswalk(db_path)

    ws.sheet_view.showGridLines = False

    # --- Header ---
    ws.merge_cells("A1:E1")
    ws["A1"] = "Deduction Code Crosswalk"
    ws["A1"].font = FONT_HEADER

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        "Maps retailer-specific deduction codes to plain-English descriptions "
        "and standardized categories. Used by the Deduction Ledger tab for code translation."
    )
    ws["A2"].font = FONT_SMALL

    # --- Column headers (row 4) ---
    header_row = 4
    header_font = Font(name=SANS, size=10, bold=True)

    for c, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.font = header_font
        cell.alignment = ALIGN_CENTER
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.freeze_panes = "A5"

    # --- Data rows ---
    for i, row_data in enumerate(rows):
        rw = header_row + 1 + i
        retailer, code, name, category, status = row_data

        ws.cell(row=rw, column=1, value=retailer.replace("_", " ").title())
        c_code = ws.cell(row=rw, column=2, value=code)
        c_code.alignment = ALIGN_CENTER
        ws.cell(row=rw, column=3, value=name)
        ws.cell(row=rw, column=4, value=category.replace("_", " ").title())
        c_status = ws.cell(row=rw, column=5, value=status)
        c_status.alignment = ALIGN_CENTER

    table_end = header_row + len(rows)

    # --- Excel Table ---
    last_col = get_column_letter(len(COLUMNS))
    table_ref = f"A{header_row}:{last_col}{table_end}"

    table = Table(displayName="tbl_CodeCrosswalk", ref=table_ref)
    table.tableStyleInfo = TABLE_STYLE
    ws.add_table(table)

    # Conditional formatting on status column
    status_range = f"E{header_row + 1}:E{table_end}"
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Verified"'], fill=FILL_GOOD),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Inferred"'], fill=FILL_WARN),
    )
